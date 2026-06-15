from datetime import date, datetime, time, timedelta
from io import BytesIO
from typing import Any
from urllib.parse import quote

import psycopg2
from fastapi import APIRouter, HTTPException, Path, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from psycopg2.extras import RealDictCursor

from db import get_connection
from schemas.equipment_maintenance import (
    EquipmentMaintenanceCreate,
    EquipmentMaintenanceRead,
    EquipmentMaintenanceUpdate,
)


router = APIRouter(prefix="/equipment-maintenance", tags=["equipment_maintenance"])
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SELECT_COLUMNS = """
    em.maintenance_id,
    em.machine_id,
    m.machine_code,
    m.machine_name,
    em.started_at,
    em.ended_at,
    em.duration_minutes,
    ROUND((em.duration_minutes::numeric / 60), 2) AS duration_hours,
    em.comment,
    em.created_at,
    em.updated_at
"""


def calculate_duration_minutes(started_at: datetime, ended_at: datetime) -> int:
    if ended_at <= started_at:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Окончание ТО должно быть позже начала.",
        )

    duration_minutes = int((ended_at - started_at).total_seconds() // 60)
    if duration_minutes <= 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Длительность ТО должна быть больше нуля.",
        )

    return duration_minutes


def ensure_machine_exists(cursor: RealDictCursor, machine_id: int) -> None:
    cursor.execute(
        """
        SELECT machine_id
        FROM machines
        WHERE machine_id = %s;
        """,
        (machine_id,),
    )
    if cursor.fetchone() is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Оборудование не найдено.",
        )


def format_print_date(value: date) -> str:
    return value.strftime("%d.%m.%Y")


def format_print_datetime(value: datetime) -> str:
    return value.strftime("%d.%m.%Y %H:%M")


def calculate_effective_maintenance_hours(
    started_at: datetime,
    ended_at: datetime,
    period_start: datetime,
    period_end_exclusive: datetime,
) -> float:
    effective_start = max(started_at, period_start)
    effective_end = min(ended_at, period_end_exclusive)
    seconds = max(0.0, (effective_end - effective_start).total_seconds())
    return round(seconds / 60 / 60, 1)


def create_maintenance_schedule_print_workbook(
    rows: list[dict[str, Any]],
    date_from: date,
    date_to: date,
    period_start: datetime,
    period_end_exclusive: datetime,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "График ТО"

    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.freeze_panes = "A7"
    worksheet.print_title_rows = "$6:$6"

    columns = [
        ("№", 5),
        ("Код оборудования", 18),
        ("Оборудование", 30),
        ("Начало ТО", 20),
        ("Окончание ТО", 20),
        ("Длительность, ч", 16),
        ("Комментарий", 45),
    ]
    for index, (_, width) in enumerate(columns, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width

    title_font = Font(size=16, bold=True, color="1E293B")
    header_font = Font(size=10, bold=True, color="0F172A")
    muted_font = Font(size=10, color="475569")
    total_font = Font(size=10, bold=True, color="0F172A")
    table_header_fill = PatternFill("solid", fgColor="DDEBEE")
    total_fill = PatternFill("solid", fgColor="F1F5F9")
    thin_side = Side(style="thin", color="B7C6CE")
    table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    worksheet.merge_cells("A1:G1")
    worksheet["A1"] = "График планового ТО оборудования"
    worksheet["A1"].font = title_font
    worksheet["A1"].alignment = Alignment(horizontal="center")

    worksheet["A2"] = f"Период: {format_print_date(date_from)} — {format_print_date(date_to)}"
    worksheet["A3"] = f"Дата печати: {format_print_datetime(datetime.now())}"
    worksheet["A2"].font = muted_font
    worksheet["A3"].font = muted_font

    table_header_row = 6
    for column_index, (title, _) in enumerate(columns, start=1):
        cell = worksheet.cell(row=table_header_row, column=column_index, value=title)
        cell.font = header_font
        cell.fill = table_header_fill
        cell.border = table_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[table_header_row].height = 28

    current_row = table_header_row
    total_hours = 0.0

    if rows:
        for index, row in enumerate(rows, start=1):
            current_row += 1
            started_at = row["started_at"]
            ended_at = row["ended_at"]
            duration_hours = calculate_effective_maintenance_hours(
                started_at=started_at,
                ended_at=ended_at,
                period_start=period_start,
                period_end_exclusive=period_end_exclusive,
            )
            total_hours += duration_hours

            row_values = [
                index,
                row.get("machine_code") or "—",
                row.get("machine_name") or "—",
                started_at,
                ended_at,
                duration_hours,
                row.get("comment") or "—",
            ]
            for column_index, value in enumerate(row_values, start=1):
                cell = worksheet.cell(row=current_row, column=column_index, value=value)
                cell.border = table_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if column_index == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                elif column_index == 6:
                    cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
                    cell.number_format = "0.0"
                elif column_index in {4, 5}:
                    cell.number_format = "DD.MM.YYYY HH:MM"
    else:
        current_row += 1
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        empty_cell = worksheet.cell(row=current_row, column=1, value="За выбранный период плановое ТО не запланировано.")
        empty_cell.font = muted_font
        empty_cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_index in range(1, 8):
            cell = worksheet.cell(row=current_row, column=column_index)
            cell.border = table_border

    current_row += 1
    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    total_label_cell = worksheet.cell(row=current_row, column=1, value=f"Итого записей: {len(rows)}")
    total_label_cell.font = total_font
    total_label_cell.fill = total_fill
    total_label_cell.alignment = Alignment(horizontal="right", vertical="center")
    hours_label_cell = worksheet.cell(row=current_row, column=6, value="Итого часов ТО:")
    hours_label_cell.font = total_font
    hours_label_cell.fill = total_fill
    hours_label_cell.alignment = Alignment(horizontal="right", vertical="center")
    hours_cell = worksheet.cell(row=current_row, column=7, value=round(total_hours, 1))
    hours_cell.font = total_font
    hours_cell.fill = total_fill
    hours_cell.alignment = Alignment(horizontal="right", vertical="center")
    hours_cell.number_format = "0.0"
    for column_index in range(1, 8):
        cell = worksheet.cell(row=current_row, column=column_index)
        cell.border = table_border
        cell.fill = total_fill

    footer_row = current_row + 4
    worksheet[f"A{footer_row}"] = "Кем сформирован график: ______________________________"
    worksheet[f"A{footer_row + 2}"] = "Подпись: ______________________________"
    worksheet[f"A{footer_row}"].font = muted_font
    worksheet[f"A{footer_row + 2}"].font = muted_font

    worksheet.row_dimensions[1].height = 24

    workbook_stream = BytesIO()
    workbook.save(workbook_stream)
    return workbook_stream.getvalue()


def build_maintenance_schedule_print_filename(date_from: date, date_to: date) -> str:
    return f"График_ТО_оборудования_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"


def require_maintenance_exists(connection, maintenance_id: int) -> dict[str, Any]:
    with connection.cursor(cursor_factory=RealDictCursor) as cursor:
        cursor.execute(
            f"""
            SELECT {SELECT_COLUMNS}
            FROM equipment_maintenance AS em
            INNER JOIN machines AS m ON m.machine_id = em.machine_id
            WHERE em.maintenance_id = %s;
            """,
            (maintenance_id,),
        )
        row = cursor.fetchone()

    if row is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Запись планового ТО не найдена.",
        )

    return row


@router.get("", response_model=list[EquipmentMaintenanceRead])
def list_equipment_maintenance(
    machine_id: int | None = Query(default=None, gt=0),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
):
    connection = None

    try:
        if date_from is not None and date_to is not None and date_from > date_to:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Дата окончания фильтра не может быть раньше даты начала.",
            )

        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            where_clauses: list[str] = []
            params: list[Any] = []

            if machine_id is not None:
                where_clauses.append("em.machine_id = %s")
                params.append(machine_id)

            if date_from is not None:
                where_clauses.append("em.ended_at >= %s")
                params.append(datetime.combine(date_from, time.min))

            if date_to is not None:
                upper_bound = datetime.combine(date_to + timedelta(days=1), time.min)
                where_clauses.append("em.started_at < %s")
                params.append(upper_bound)

            where_sql = ""
            if where_clauses:
                where_sql = "WHERE " + " AND ".join(where_clauses)

            has_filters = bool(where_clauses)
            limit_sql = "" if has_filters else "LIMIT 100"

            cursor.execute(
                f"""
                SELECT {SELECT_COLUMNS}
                FROM equipment_maintenance AS em
                INNER JOIN machines AS m ON m.machine_id = em.machine_id
                {where_sql}
                ORDER BY em.started_at DESC, em.maintenance_id DESC
                {limit_sql};
                """,
                tuple(params),
            )
            rows = cursor.fetchall()

        return rows
    except HTTPException:
        raise
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось получить список планового ТО.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.get("/print")
def print_equipment_maintenance_schedule(
    date_from: date = Query(...),
    date_to: date = Query(...),
):
    connection = None

    try:
        if date_from > date_to:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Дата окончания периода печати не может быть раньше даты начала.",
            )

        period_start = datetime.combine(date_from, time.min)
        period_end_exclusive = datetime.combine(date_to + timedelta(days=1), time.min)

        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                f"""
                SELECT {SELECT_COLUMNS}
                FROM equipment_maintenance AS em
                INNER JOIN machines AS m ON m.machine_id = em.machine_id
                WHERE em.started_at < %s
                  AND em.ended_at > %s
                ORDER BY em.started_at ASC, m.machine_code ASC, em.ended_at ASC;
                """,
                (period_end_exclusive, period_start),
            )
            rows = cursor.fetchall()

        workbook_bytes = create_maintenance_schedule_print_workbook(
            rows=list(rows),
            date_from=date_from,
            date_to=date_to,
            period_start=period_start,
            period_end_exclusive=period_end_exclusive,
        )
        filename = build_maintenance_schedule_print_filename(date_from, date_to)
        encoded_filename = quote(filename)
        return StreamingResponse(
            BytesIO(workbook_bytes),
            media_type=XLSX_MEDIA_TYPE,
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            },
        )
    except HTTPException:
        raise
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось сформировать печатную форму графика ТО.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.post("", response_model=EquipmentMaintenanceRead, status_code=status.HTTP_201_CREATED)
def create_equipment_maintenance(payload: EquipmentMaintenanceCreate):
    connection = None

    try:
        duration_minutes = calculate_duration_minutes(payload.started_at, payload.ended_at)

        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            ensure_machine_exists(cursor, payload.machine_id)

            # NOTE: Пересекающиеся интервалы ТО по одному оборудованию разрешены.
            # На этапе расчёта доступности интервалы нужно объединять, чтобы не задваивать простой.
            cursor.execute(
                """
                INSERT INTO equipment_maintenance (
                    machine_id,
                    started_at,
                    ended_at,
                    duration_minutes,
                    comment
                )
                VALUES (%s, %s, %s, %s, %s)
                RETURNING maintenance_id;
                """,
                (
                    payload.machine_id,
                    payload.started_at,
                    payload.ended_at,
                    duration_minutes,
                    payload.comment,
                ),
            )
            created_row = cursor.fetchone()

        connection.commit()
        return require_maintenance_exists(connection, int(created_row["maintenance_id"]))
    except HTTPException:
        if connection is not None:
            connection.rollback()
        raise
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось создать запись планового ТО.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.put("/{maintenance_id}", response_model=EquipmentMaintenanceRead)
def update_equipment_maintenance(
    payload: EquipmentMaintenanceUpdate,
    maintenance_id: int = Path(..., gt=0),
):
    connection = None

    try:
        duration_minutes = calculate_duration_minutes(payload.started_at, payload.ended_at)

        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                SELECT maintenance_id
                FROM equipment_maintenance
                WHERE maintenance_id = %s
                FOR UPDATE;
                """,
                (maintenance_id,),
            )
            if cursor.fetchone() is None:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Запись планового ТО не найдена.",
                )

            ensure_machine_exists(cursor, payload.machine_id)

            cursor.execute(
                """
                UPDATE equipment_maintenance
                SET
                    machine_id = %s,
                    started_at = %s,
                    ended_at = %s,
                    duration_minutes = %s,
                    comment = %s,
                    updated_at = NOW()
                WHERE maintenance_id = %s;
                """,
                (
                    payload.machine_id,
                    payload.started_at,
                    payload.ended_at,
                    duration_minutes,
                    payload.comment,
                    maintenance_id,
                ),
            )

        connection.commit()
        return require_maintenance_exists(connection, maintenance_id)
    except HTTPException:
        if connection is not None:
            connection.rollback()
        raise
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось обновить запись планового ТО.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.delete("/{maintenance_id}")
def delete_equipment_maintenance(maintenance_id: int = Path(..., gt=0)):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                DELETE FROM equipment_maintenance
                WHERE maintenance_id = %s
                RETURNING maintenance_id;
                """,
                (maintenance_id,),
            )
            deleted_row = cursor.fetchone()

        if deleted_row is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Запись планового ТО не найдена.",
            )

        connection.commit()
        return {
            "maintenance_id": int(deleted_row["maintenance_id"]),
            "message": "Запись планового ТО удалена.",
        }
    except HTTPException:
        if connection is not None:
            connection.rollback()
        raise
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось удалить запись планового ТО.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()
