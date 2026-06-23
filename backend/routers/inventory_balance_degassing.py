from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re
from typing import Any

import psycopg2
from fastapi import APIRouter, Depends, File, HTTPException, Path, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel
from psycopg2.extras import RealDictCursor

from auth.rbac import require_roles
from db import get_connection
from schemas.inventory_balance_degassing import (
    InventoryBalanceDegassingCreate,
    InventoryBalanceDegassingDeleteResponse,
    InventoryBalanceDegassingImportResponse,
    InventoryBalanceDegassingRead,
    InventoryBalanceDegassingSuggestionItem,
    InventoryBalanceDegassingSuggestionReportResponse,
    InventoryBalanceDegassingSuggestionSourceBatch,
    InventoryBalanceDegassingUpdate,
)


router = APIRouter(prefix="/inventory-balance-degassing", tags=["inventory_balance_degassing"])

INVENTORY_BALANCE_DEGASSING_READ_ROLES = ("planner", "viewer")
INVENTORY_BALANCE_DEGASSING_WRITE_ROLES = ("planner",)

SELECT_COLUMNS = """
    ibd.balance_degassing_id,
    ibd.as_of_date,
    ibd.nomenclature_id,
    n.nomenclature_code,
    n.nomenclature_name,
    n.unit_of_measure,
    ibd.qty,
    ibd.available_at,
    ibd.comment,
    ibd.created_at,
    ibd.updated_at
"""

HEADER_FIELD_AS_OF_DATE = "as_of_date"
HEADER_FIELD_NOMENCLATURE_CODE = "nomenclature_code"
HEADER_FIELD_QTY = "qty"
HEADER_FIELD_AVAILABLE_AT = "available_at"
HEADER_FIELD_COMMENT = "comment"

REQUIRED_IMPORT_HEADERS = {
    HEADER_FIELD_AS_OF_DATE,
    HEADER_FIELD_NOMENCLATURE_CODE,
    HEADER_FIELD_QTY,
    HEADER_FIELD_AVAILABLE_AT,
}

HEADER_ALIASES: dict[str, set[str]] = {
    HEADER_FIELD_AS_OF_DATE: {"датаостатков", "датаостатка", "asofdate"},
    HEADER_FIELD_NOMENCLATURE_CODE: {"кодноменклатуры", "nomenclaturecode"},
    HEADER_FIELD_QTY: {"количество", "qty"},
    HEADER_FIELD_AVAILABLE_AT: {"доступнос", "датадоступности", "availableat"},
    HEADER_FIELD_COMMENT: {"комментарий", "comment"},
}

ALLOWED_AVAILABLE_TIMES = {"07:00", "19:00"}


def get_balance_start_datetime(as_of_date: date) -> datetime:
    return datetime.combine(as_of_date, time(hour=7))


def validate_qty(qty: Decimal) -> None:
    if qty <= 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Количество должно быть больше 0.",
        )


def validate_available_at(as_of_date: date, available_at: datetime) -> None:
    if available_at <= get_balance_start_datetime(as_of_date):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Дата доступности должна быть позже даты остатков 07:00.",
        )


def normalize_header_name(value: object) -> str:
    if value is None:
        return ""

    normalized_value = str(value).strip().lower().replace("ё", "е")
    return re.sub(r"[\s_\-./\\]+", "", normalized_value)


def resolve_import_header(value: object) -> str | None:
    normalized_value = normalize_header_name(value)
    if not normalized_value:
        return None

    for field_name, aliases in HEADER_ALIASES.items():
        if normalized_value in aliases:
            return field_name

    return None


def normalize_code(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_comment(value: object) -> str | None:
    if value is None:
        return None

    normalized_value = str(value).strip()
    return normalized_value or None


def normalize_qty(value: object) -> tuple[Decimal | None, str | None]:
    if value is None:
        return None, "количество должно быть больше 0."

    if isinstance(value, Decimal):
        decimal_value = value
    elif isinstance(value, (int, float)):
        decimal_value = Decimal(str(value))
    else:
        raw_text = str(value).strip().replace(" ", "")
        if not raw_text:
            return None, "количество должно быть больше 0."

        raw_text = raw_text.replace(",", ".")
        try:
            decimal_value = Decimal(raw_text)
        except InvalidOperation:
            return None, "количество должно быть больше 0."

    if decimal_value <= 0:
        return None, "количество должно быть больше 0."

    return decimal_value.quantize(Decimal("0.001")), None


def normalize_as_of_date(value: object) -> tuple[date | None, str | None]:
    if value is None:
        return None, "не заполнена дата остатков."

    if isinstance(value, datetime):
        return value.date(), None

    if isinstance(value, date):
        return value, None

    if isinstance(value, (int, float)):
        try:
            excel_datetime = from_excel(value)
            if isinstance(excel_datetime, datetime):
                return excel_datetime.date(), None
            if isinstance(excel_datetime, date):
                return excel_datetime, None
        except Exception:
            return None, "не заполнена дата остатков."

        return None, "не заполнена дата остатков."

    raw_value = str(value).strip()
    if not raw_value:
        return None, "не заполнена дата остатков."

    date_formats = [
        "%Y-%m-%d",
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d-%m-%Y",
    ]
    for date_format in date_formats:
        try:
            return datetime.strptime(raw_value, date_format).date(), None
        except ValueError:
            continue

    return None, "не заполнена дата остатков."


def normalize_available_at_import(value: object) -> tuple[datetime | None, str | None]:
    if value is None:
        return None, "не заполнена дата доступности."

    if isinstance(value, datetime):
        return value.replace(second=0, microsecond=0), None

    if isinstance(value, date):
        return datetime.combine(value, time.min), None

    if isinstance(value, (int, float)):
        try:
            excel_datetime = from_excel(value)
            if isinstance(excel_datetime, datetime):
                return excel_datetime.replace(second=0, microsecond=0), None
            if isinstance(excel_datetime, date):
                return datetime.combine(excel_datetime, time.min), None
        except Exception:
            return None, "не заполнена дата доступности."

        return None, "не заполнена дата доступности."

    raw_value = str(value).strip()
    if not raw_value:
        return None, "не заполнена дата доступности."

    datetime_formats = [
        "%d.%m.%Y %H:%M",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M",
        "%Y/%m/%d %H:%M",
        "%d-%m-%Y %H:%M",
        "%Y-%m-%dT%H:%M",
        "%d.%m.%Y",
        "%Y-%m-%d",
    ]
    for datetime_format in datetime_formats:
        try:
            parsed_value = datetime.strptime(raw_value, datetime_format)
            return parsed_value.replace(second=0, microsecond=0), None
        except ValueError:
            continue

    return None, "не заполнена дата доступности."


def validate_import_file(file: UploadFile, file_bytes: bytes) -> None:
    file_name = file.filename or ""
    if not file_name.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Поддерживается только формат .xlsx.",
        )

    if len(file_bytes) == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Файл пустой.",
        )


def read_import_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Не удалось прочитать Excel-файл. Проверьте формат .xlsx.",
        ) from exc

    try:
        sheet = workbook.active
        max_column = sheet.max_column or 0
        max_row = sheet.max_row or 0

        if max_column == 0 or max_row == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Файл пустой.",
            )

        header_indexes: dict[str, int] = {}
        for column_index in range(1, max_column + 1):
            header_name = resolve_import_header(sheet.cell(row=1, column=column_index).value)
            if header_name and header_name not in header_indexes:
                header_indexes[header_name] = column_index

        missing_headers = [
            header_name
            for header_name in REQUIRED_IMPORT_HEADERS
            if header_name not in header_indexes
        ]
        if missing_headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "Не найдены обязательные колонки шаблона: "
                    "Дата остатков, Код номенклатуры, Количество, Доступно с."
                ),
            )

        rows: list[dict[str, Any]] = []
        for row_index in range(2, max_row + 1):
            raw_as_of_date = sheet.cell(row=row_index, column=header_indexes[HEADER_FIELD_AS_OF_DATE]).value
            raw_code = sheet.cell(row=row_index, column=header_indexes[HEADER_FIELD_NOMENCLATURE_CODE]).value
            raw_qty = sheet.cell(row=row_index, column=header_indexes[HEADER_FIELD_QTY]).value
            raw_available_at = sheet.cell(row=row_index, column=header_indexes[HEADER_FIELD_AVAILABLE_AT]).value
            raw_comment = None
            if HEADER_FIELD_COMMENT in header_indexes:
                raw_comment = sheet.cell(row=row_index, column=header_indexes[HEADER_FIELD_COMMENT]).value

            is_empty_row = (
                (raw_as_of_date is None or str(raw_as_of_date).strip() == "")
                and normalize_code(raw_code) == ""
                and (raw_qty is None or str(raw_qty).strip() == "")
                and (raw_available_at is None or str(raw_available_at).strip() == "")
                and (raw_comment is None or str(raw_comment).strip() == "")
            )
            if is_empty_row:
                continue

            normalized_as_of_date, as_of_date_error = normalize_as_of_date(raw_as_of_date)
            normalized_code = normalize_code(raw_code)
            normalized_qty, qty_error = normalize_qty(raw_qty)
            normalized_available_at, available_at_error = normalize_available_at_import(raw_available_at)

            rows.append(
                {
                    "row_no": row_index,
                    "as_of_date": normalized_as_of_date,
                    "nomenclature_code": normalized_code or None,
                    "nomenclature_code_key": normalized_code.upper() if normalized_code else None,
                    "qty": normalized_qty,
                    "available_at": normalized_available_at,
                    "comment": normalize_comment(raw_comment),
                    "errors": [
                        error
                        for error in [
                            as_of_date_error,
                            None if normalized_code else "не заполнен код номенклатуры.",
                            qty_error,
                            available_at_error,
                        ]
                        if error
                    ],
                }
            )

        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Файл пустой.",
            )

        return rows
    finally:
        workbook.close()


def get_available_time_label(value: datetime) -> str:
    return value.strftime("%H:%M")


def fetch_nomenclature_by_code(cursor: RealDictCursor) -> dict[str, dict[str, Any]]:
    cursor.execute(
        """
        SELECT nomenclature_id, nomenclature_code
        FROM nomenclature;
        """
    )
    rows = cursor.fetchall()
    return {str(row["nomenclature_code"]).strip().upper(): row for row in rows}


def fetch_inventory_balances_by_groups(
    cursor: RealDictCursor,
    as_of_dates: set[date],
    nomenclature_ids: set[int],
) -> dict[tuple[date, int], Decimal]:
    if not as_of_dates or not nomenclature_ids:
        return {}

    cursor.execute(
        """
        SELECT as_of_date, nomenclature_id, available_qty
        FROM inventory_balance
        WHERE as_of_date = ANY(%s)
          AND nomenclature_id = ANY(%s);
        """,
        (list(as_of_dates), list(nomenclature_ids)),
    )
    rows = cursor.fetchall()
    return {
        (row["as_of_date"], row["nomenclature_id"]): row["available_qty"]
        for row in rows
    }


def create_template_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Остатки в дегазации"
    sheet.append(
        [
            "Дата остатков",
            "Код номенклатуры",
            "Количество",
            "Доступно с",
            "Комментарий",
        ]
    )
    sheet.append(
        [
            "01.06.2026",
            "NM-021",
            5000,
            "04.06.2026 19:00",
            "Выпуск последней недели месяца",
        ]
    )

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def format_import_date(value: date) -> str:
    return value.strftime("%d.%m.%Y")


def format_import_datetime(value: datetime) -> str:
    return value.strftime("%d.%m.%Y %H:%M")


def get_suggestion_report_check_at(as_of_date: date) -> datetime:
    return get_balance_start_datetime(as_of_date)


def get_shift_label(shift_type: str) -> str:
    return "день" if shift_type == "day" else "ночь"


def fetch_suggestion_report_items(
    cursor: RealDictCursor,
    as_of_date: date,
    lookback_days: int,
) -> tuple[datetime, list[InventoryBalanceDegassingSuggestionItem]]:
    check_at = get_suggestion_report_check_at(as_of_date)
    date_from = as_of_date - timedelta(days=lookback_days)
    date_to = as_of_date

    cursor.execute(
        """
        WITH wait_hours_by_nomenclature AS (
            SELECT DISTINCT ON (r.result_nomenclature_id)
                r.result_nomenclature_id AS nomenclature_id,
                COALESCE(step_row.post_process_wait_hours, 0) AS post_process_wait_hours
            FROM routes AS r
            INNER JOIN LATERAL (
                SELECT rs.post_process_wait_hours
                FROM route_steps AS rs
                WHERE rs.route_id = r.route_id
                  AND rs.output_nomenclature_id = r.result_nomenclature_id
                ORDER BY rs.step_no DESC, rs.route_step_id DESC
                LIMIT 1
            ) AS step_row ON TRUE
            WHERE r.is_active = TRUE
            ORDER BY
                r.result_nomenclature_id ASC,
                r.route_id ASC
        )
        SELECT
            pa.production_actual_id,
            pa.nomenclature_id,
            n.nomenclature_code,
            n.nomenclature_name,
            COALESCE(NULLIF(BTRIM(pa.unit_of_measure), ''), n.unit_of_measure, '') AS unit_of_measure,
            pa.actual_date,
            pa.shift_type,
            COALESCE(pa.actual_qty, 0) AS actual_qty,
            CASE
                WHEN pa.shift_type = 'day'
                    THEN pa.actual_date::timestamp + INTERVAL '19 hour'
                ELSE pa.actual_date::timestamp + INTERVAL '1 day' + INTERVAL '7 hour'
            END AS shift_finish_at,
            COALESCE(rs.post_process_wait_hours, wait_hours.post_process_wait_hours, 0) AS degassing_hours,
            (
                CASE
                    WHEN pa.shift_type = 'day'
                        THEN pa.actual_date::timestamp + INTERVAL '19 hour'
                    ELSE pa.actual_date::timestamp + INTERVAL '1 day' + INTERVAL '7 hour'
                END
                + (COALESCE(rs.post_process_wait_hours, wait_hours.post_process_wait_hours, 0) * INTERVAL '1 hour')
            ) AS available_at
        FROM production_actuals AS pa
        INNER JOIN nomenclature AS n
            ON n.nomenclature_id = pa.nomenclature_id
        LEFT JOIN production_week_lines AS pwl
            ON pwl.production_week_line_id = pa.production_week_line_id
        LEFT JOIN route_step_equipment AS rse
            ON rse.step_equipment_id = pwl.route_step_equipment_id
        LEFT JOIN route_steps AS rs
            ON rs.route_step_id = rse.route_step_id
        LEFT JOIN wait_hours_by_nomenclature AS wait_hours
            ON wait_hours.nomenclature_id = pa.nomenclature_id
        WHERE pa.actual_date >= %s
          AND pa.actual_date < %s
          AND COALESCE(rs.post_process_wait_hours, wait_hours.post_process_wait_hours, 0) > 0
        ORDER BY
            n.nomenclature_code ASC,
            available_at ASC,
            pa.actual_date ASC,
            pa.production_actual_id ASC;
        """,
        (date_from, date_to),
    )

    grouped_items: dict[tuple[int, datetime], dict[str, Any]] = {}
    nomenclature_ids: set[int] = set()

    for row in cursor.fetchall():
        available_at = row["available_at"]
        actual_qty = row["actual_qty"]
        if not isinstance(available_at, datetime):
            continue

        if not isinstance(actual_qty, Decimal):
            actual_qty = Decimal(str(actual_qty or 0))

        if actual_qty <= Decimal("0") or available_at <= check_at:
            continue

        nomenclature_id = int(row["nomenclature_id"])
        key = (nomenclature_id, available_at)
        nomenclature_ids.add(nomenclature_id)

        source_batch = InventoryBalanceDegassingSuggestionSourceBatch(
            actual_date=row["actual_date"],
            shift_type=row["shift_type"],
            actual_qty=actual_qty,
        )

        if key not in grouped_items:
            grouped_items[key] = {
                "nomenclature_id": nomenclature_id,
                "nomenclature_code": str(row["nomenclature_code"] or ""),
                "nomenclature_name": str(row["nomenclature_name"] or ""),
                "unit_of_measure": str(row["unit_of_measure"] or ""),
                "actual_qty": Decimal("0"),
                "available_at": available_at,
                "degassing_hours": Decimal(str(row["degassing_hours"] or 0)),
                "source_batches": [],
                "actual_date": row["actual_date"],
                "shift_type": row["shift_type"],
                "shift_finish_at": row["shift_finish_at"],
            }

        grouped_items[key]["actual_qty"] += actual_qty
        grouped_items[key]["source_batches"].append(source_batch)

        if len(grouped_items[key]["source_batches"]) > 1:
            grouped_items[key]["actual_date"] = None
            grouped_items[key]["shift_type"] = None
            grouped_items[key]["shift_finish_at"] = None

    inventory_balances = fetch_inventory_balances_by_groups(
        cursor,
        {as_of_date},
        nomenclature_ids,
    )

    items: list[InventoryBalanceDegassingSuggestionItem] = []
    for key in sorted(
        grouped_items,
        key=lambda item_key: (
            grouped_items[item_key]["nomenclature_code"],
            item_key[1],
        ),
    ):
        item = grouped_items[key]
        balance_key = (as_of_date, item["nomenclature_id"])
        inventory_balance_qty = inventory_balances.get(balance_key)
        source_batches = item["source_batches"]
        if len(source_batches) == 1:
            only_batch = source_batches[0]
            source_summary = f"{format_import_date(only_batch.actual_date)} / {get_shift_label(only_batch.shift_type)}"
        else:
            source_summary = "несколько фактов"

        items.append(
            InventoryBalanceDegassingSuggestionItem(
                nomenclature_id=item["nomenclature_id"],
                nomenclature_code=item["nomenclature_code"],
                nomenclature_name=item["nomenclature_name"],
                unit_of_measure=item["unit_of_measure"],
                actual_date=item["actual_date"],
                shift_type=item["shift_type"],
                actual_qty=item["actual_qty"].quantize(Decimal("0.001")),
                shift_finish_at=item["shift_finish_at"],
                degassing_hours=item["degassing_hours"].quantize(Decimal("0.001")),
                available_at=item["available_at"],
                status="Будет доступен",
                has_inventory_balance=inventory_balance_qty is not None,
                inventory_balance_qty=(
                    inventory_balance_qty.quantize(Decimal("0.001"))
                    if isinstance(inventory_balance_qty, Decimal)
                    else inventory_balance_qty
                ),
                source_summary=source_summary,
                source_batches=source_batches,
            )
        )

    return check_at, items


def create_suggestion_report_export_workbook(
    as_of_date: date,
    items: list[InventoryBalanceDegassingSuggestionItem],
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ПФ в дегазации"
    sheet.append(
        [
            "Дата остатков",
            "Код номенклатуры",
            "Количество",
            "Доступно с",
            "Комментарий",
        ]
    )

    for item in items:
        sheet.append(
            [
                format_import_date(as_of_date),
                item.nomenclature_code,
                float(item.actual_qty),
                format_import_datetime(item.available_at),
                "Выпуск последней недели месяца",
            ]
        )

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def ensure_nomenclature_exists(cursor: RealDictCursor, nomenclature_id: int) -> None:
    cursor.execute(
        """
        SELECT nomenclature_id
        FROM nomenclature
        WHERE nomenclature_id = %s;
        """,
        (nomenclature_id,),
    )
    if cursor.fetchone() is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Номенклатура не найдена.",
        )


def lock_inventory_balance(cursor: RealDictCursor, as_of_date: date, nomenclature_id: int) -> Decimal:
    cursor.execute(
        """
        SELECT balance_id, available_qty
        FROM inventory_balance
        WHERE as_of_date = %s
          AND nomenclature_id = %s
        FOR UPDATE;
        """,
        (as_of_date, nomenclature_id),
    )
    row = cursor.fetchone()
    if row is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Для выбранной даты остатков и номенклатуры не найден общий остаток.",
        )
    return row["available_qty"]


def get_existing_degassing_qty(
    cursor: RealDictCursor,
    as_of_date: date,
    nomenclature_id: int,
    exclude_balance_degassing_id: int | None = None,
) -> Decimal:
    where_sql = ""
    params: list[Any] = [as_of_date, nomenclature_id]

    if exclude_balance_degassing_id is not None:
        where_sql = "AND balance_degassing_id <> %s"
        params.append(exclude_balance_degassing_id)

    cursor.execute(
        f"""
        SELECT COALESCE(SUM(qty), 0) AS degassing_qty
        FROM inventory_balance_degassing
        WHERE as_of_date = %s
          AND nomenclature_id = %s
          {where_sql};
        """,
        tuple(params),
    )
    row = cursor.fetchone()
    return row["degassing_qty"] if row is not None else Decimal("0")


def ensure_degassing_qty_within_balance(
    cursor: RealDictCursor,
    as_of_date: date,
    nomenclature_id: int,
    qty: Decimal,
    exclude_balance_degassing_id: int | None = None,
) -> None:
    available_qty = lock_inventory_balance(cursor, as_of_date, nomenclature_id)
    existing_degassing_qty = get_existing_degassing_qty(
        cursor,
        as_of_date,
        nomenclature_id,
        exclude_balance_degassing_id=exclude_balance_degassing_id,
    )

    if existing_degassing_qty + qty > available_qty:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Количество остатков в дегазации превышает общий остаток на дату остатков.",
        )


def require_inventory_balance_degassing_exists(connection, balance_degassing_id: int) -> dict[str, Any]:
    with connection.cursor(cursor_factory=RealDictCursor) as cursor:
        cursor.execute(
            f"""
            SELECT {SELECT_COLUMNS}
            FROM inventory_balance_degassing AS ibd
            INNER JOIN nomenclature AS n ON n.nomenclature_id = ibd.nomenclature_id
            WHERE ibd.balance_degassing_id = %s;
            """,
            (balance_degassing_id,),
        )
        row = cursor.fetchone()

    if row is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Запись остатков в дегазации не найдена.",
        )

    return row


@router.get("", response_model=list[InventoryBalanceDegassingRead], dependencies=[Depends(require_roles(*INVENTORY_BALANCE_DEGASSING_READ_ROLES))])
def list_inventory_balance_degassing(
    as_of_date: date | None = Query(default=None),
    nomenclature_id: int | None = Query(default=None, gt=0),
):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            where_clauses: list[str] = []
            params: list[Any] = []

            if as_of_date is not None:
                where_clauses.append("ibd.as_of_date = %s")
                params.append(as_of_date)

            if nomenclature_id is not None:
                where_clauses.append("ibd.nomenclature_id = %s")
                params.append(nomenclature_id)

            where_sql = ""
            if where_clauses:
                where_sql = "WHERE " + " AND ".join(where_clauses)

            cursor.execute(
                f"""
                SELECT {SELECT_COLUMNS}
                FROM inventory_balance_degassing AS ibd
                INNER JOIN nomenclature AS n ON n.nomenclature_id = ibd.nomenclature_id
                {where_sql}
                ORDER BY ibd.as_of_date DESC, n.nomenclature_code ASC, ibd.available_at ASC;
                """,
                tuple(params),
            )
            return cursor.fetchall()
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось получить список остатков в дегазации.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.get("/suggestion-report", response_model=InventoryBalanceDegassingSuggestionReportResponse, dependencies=[Depends(require_roles(*INVENTORY_BALANCE_DEGASSING_READ_ROLES))])
def get_inventory_balance_degassing_suggestion_report(
    as_of_date: date = Query(...),
    lookback_days: int = Query(default=7, ge=1),
):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            check_at, items = fetch_suggestion_report_items(cursor, as_of_date, lookback_days)
            return InventoryBalanceDegassingSuggestionReportResponse(
                as_of_date=as_of_date,
                check_at=check_at,
                lookback_days=lookback_days,
                items=items,
            )
    except HTTPException:
        raise
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось сформировать отчёт по ПФ в дегазации.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.get("/suggestion-report/export", dependencies=[Depends(require_roles(*INVENTORY_BALANCE_DEGASSING_READ_ROLES))])
def export_inventory_balance_degassing_suggestion_report(
    as_of_date: date = Query(...),
    lookback_days: int = Query(default=7, ge=1),
):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            _, items = fetch_suggestion_report_items(cursor, as_of_date, lookback_days)

        workbook_bytes = create_suggestion_report_export_workbook(as_of_date, items)
        return StreamingResponse(
            BytesIO(workbook_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": (
                    f'attachment; filename="inventory_balance_degassing_suggestion_{as_of_date.isoformat()}.xlsx"'
                ),
            },
        )
    except HTTPException:
        raise
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось выгрузить отчёт по ПФ в дегазации.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.get("/template", dependencies=[Depends(require_roles(*INVENTORY_BALANCE_DEGASSING_READ_ROLES))])
def download_inventory_balance_degassing_template():
    template_content = create_template_workbook()
    return StreamingResponse(
        BytesIO(template_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="inventory_balance_degassing_template.xlsx"',
        },
    )


@router.post("/import", response_model=InventoryBalanceDegassingImportResponse, dependencies=[Depends(require_roles(*INVENTORY_BALANCE_DEGASSING_WRITE_ROLES))])
async def import_inventory_balance_degassing(file: UploadFile = File(...)):
    file_bytes = await file.read()
    validate_import_file(file, file_bytes)
    parsed_rows = read_import_rows(file_bytes)

    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            nomenclature_by_code = fetch_nomenclature_by_code(cursor)

            normalized_rows: list[dict[str, Any]] = []
            validation_errors: list[str] = []
            as_of_dates: set[date] = set()
            nomenclature_ids: set[int] = set()
            grouped_qty: dict[tuple[date, int], Decimal] = {}
            grouped_codes: dict[tuple[date, int], str] = {}

            for parsed_row in parsed_rows:
                row_no = parsed_row["row_no"]
                row_errors = [f"Строка {row_no}: {message}" for message in parsed_row["errors"]]

                nomenclature_key = parsed_row["nomenclature_code_key"]
                nomenclature_row = nomenclature_by_code.get(nomenclature_key) if nomenclature_key else None
                if parsed_row["nomenclature_code"] and nomenclature_row is None:
                    row_errors.append(
                        f"Строка {row_no}: номенклатура с кодом {parsed_row['nomenclature_code']} не найдена."
                    )

                available_at = parsed_row["available_at"]
                if available_at is not None and get_available_time_label(available_at) not in ALLOWED_AVAILABLE_TIMES:
                    row_errors.append(f"Строка {row_no}: время доступности должно быть 07:00 или 19:00.")

                as_of_date = parsed_row["as_of_date"]
                if as_of_date is not None and available_at is not None and available_at <= get_balance_start_datetime(as_of_date):
                    row_errors.append(
                        f"Строка {row_no}: дата доступности должна быть позже даты остатков 07:00."
                    )

                if row_errors:
                    validation_errors.extend(row_errors)
                    continue

                nomenclature_id = int(nomenclature_row["nomenclature_id"])
                qty = parsed_row["qty"]
                row_data = {
                    "row_no": row_no,
                    "as_of_date": as_of_date,
                    "nomenclature_id": nomenclature_id,
                    "nomenclature_code": parsed_row["nomenclature_code"],
                    "qty": qty,
                    "available_at": available_at,
                    "comment": parsed_row["comment"],
                }
                normalized_rows.append(row_data)
                as_of_dates.add(as_of_date)
                nomenclature_ids.add(nomenclature_id)

                group_key = (as_of_date, nomenclature_id)
                grouped_qty[group_key] = grouped_qty.get(group_key, Decimal("0")) + qty
                grouped_codes[group_key] = parsed_row["nomenclature_code"]

            inventory_balances = fetch_inventory_balances_by_groups(cursor, as_of_dates, nomenclature_ids)

            for row in normalized_rows:
                group_key = (row["as_of_date"], row["nomenclature_id"])
                if group_key not in inventory_balances:
                    validation_errors.append(
                        f"Строка {row['row_no']}: для даты остатков и номенклатуры не найден общий остаток."
                    )

            for group_key, import_qty in grouped_qty.items():
                available_qty = inventory_balances.get(group_key)
                if available_qty is None:
                    continue

                if import_qty > available_qty:
                    as_of_date, nomenclature_id = group_key
                    _ = nomenclature_id
                    validation_errors.append(
                        "Сумма остатков в дегазации по "
                        f"{grouped_codes[group_key]} на {as_of_date.strftime('%d.%m.%Y')} превышает общий остаток."
                    )

            if validation_errors:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=validation_errors,
                )

            if as_of_dates:
                cursor.execute(
                    """
                    DELETE FROM inventory_balance_degassing
                    WHERE as_of_date = ANY(%s);
                    """,
                    (list(as_of_dates),),
                )

            for row in normalized_rows:
                cursor.execute(
                    """
                    INSERT INTO inventory_balance_degassing (
                        as_of_date,
                        nomenclature_id,
                        qty,
                        available_at,
                        comment
                    )
                    VALUES (%s, %s, %s, %s, %s);
                    """,
                    (
                        row["as_of_date"],
                        row["nomenclature_id"],
                        row["qty"],
                        row["available_at"],
                        row["comment"],
                    ),
                )

        connection.commit()
        return InventoryBalanceDegassingImportResponse(
            imported_count=len(normalized_rows),
            affected_dates=sorted(as_of_dates),
            message="Остатки в дегазации загружены.",
        )
    except HTTPException:
        if connection is not None:
            connection.rollback()
        raise
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось загрузить остатки в дегазации.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.post("", response_model=InventoryBalanceDegassingRead, status_code=status.HTTP_201_CREATED, dependencies=[Depends(require_roles(*INVENTORY_BALANCE_DEGASSING_WRITE_ROLES))])
def create_inventory_balance_degassing(payload: InventoryBalanceDegassingCreate):
    connection = None

    try:
        validate_qty(payload.qty)
        validate_available_at(payload.as_of_date, payload.available_at)

        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            ensure_nomenclature_exists(cursor, payload.nomenclature_id)
            ensure_degassing_qty_within_balance(
                cursor,
                payload.as_of_date,
                payload.nomenclature_id,
                payload.qty,
            )

            cursor.execute(
                """
                INSERT INTO inventory_balance_degassing (
                    as_of_date,
                    nomenclature_id,
                    qty,
                    available_at,
                    comment
                )
                VALUES (%s, %s, %s, %s, %s)
                RETURNING balance_degassing_id;
                """,
                (
                    payload.as_of_date,
                    payload.nomenclature_id,
                    payload.qty,
                    payload.available_at,
                    payload.comment,
                ),
            )
            created_row = cursor.fetchone()

        connection.commit()
        return require_inventory_balance_degassing_exists(connection, int(created_row["balance_degassing_id"]))
    except HTTPException:
        if connection is not None:
            connection.rollback()
        raise
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось создать запись остатков в дегазации.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.put("/{balance_degassing_id}", response_model=InventoryBalanceDegassingRead, dependencies=[Depends(require_roles(*INVENTORY_BALANCE_DEGASSING_WRITE_ROLES))])
def update_inventory_balance_degassing(
    payload: InventoryBalanceDegassingUpdate,
    balance_degassing_id: int = Path(..., gt=0),
):
    connection = None

    try:
        validate_qty(payload.qty)

        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                SELECT balance_degassing_id, as_of_date, nomenclature_id
                FROM inventory_balance_degassing
                WHERE balance_degassing_id = %s
                FOR UPDATE;
                """,
                (balance_degassing_id,),
            )
            current_row = cursor.fetchone()
            if current_row is None:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Запись остатков в дегазации не найдена.",
                )

            validate_available_at(current_row["as_of_date"], payload.available_at)
            ensure_nomenclature_exists(cursor, current_row["nomenclature_id"])
            ensure_degassing_qty_within_balance(
                cursor,
                current_row["as_of_date"],
                current_row["nomenclature_id"],
                payload.qty,
                exclude_balance_degassing_id=balance_degassing_id,
            )

            cursor.execute(
                """
                UPDATE inventory_balance_degassing
                SET
                    qty = %s,
                    available_at = %s,
                    comment = %s,
                    updated_at = NOW()
                WHERE balance_degassing_id = %s;
                """,
                (
                    payload.qty,
                    payload.available_at,
                    payload.comment,
                    balance_degassing_id,
                ),
            )

        connection.commit()
        return require_inventory_balance_degassing_exists(connection, balance_degassing_id)
    except HTTPException:
        if connection is not None:
            connection.rollback()
        raise
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось обновить запись остатков в дегазации.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.delete("/{balance_degassing_id}", response_model=InventoryBalanceDegassingDeleteResponse, dependencies=[Depends(require_roles(*INVENTORY_BALANCE_DEGASSING_WRITE_ROLES))])
def delete_inventory_balance_degassing(balance_degassing_id: int = Path(..., gt=0)):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                DELETE FROM inventory_balance_degassing
                WHERE balance_degassing_id = %s
                RETURNING balance_degassing_id;
                """,
                (balance_degassing_id,),
            )
            deleted_row = cursor.fetchone()
            if deleted_row is None:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Запись остатков в дегазации не найдена.",
                )

        connection.commit()
        return InventoryBalanceDegassingDeleteResponse(
            balance_degassing_id=balance_degassing_id,
            message="Запись остатков в дегазации удалена.",
        )
    except HTTPException:
        if connection is not None:
            connection.rollback()
        raise
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось удалить запись остатков в дегазации.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()
