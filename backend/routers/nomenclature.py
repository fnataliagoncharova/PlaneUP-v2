from typing import List
from io import BytesIO
import re

import psycopg2
from fastapi import APIRouter, Depends, File, Form, HTTPException, Path, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from psycopg2.errors import CheckViolation, UniqueViolation
from psycopg2.extras import RealDictCursor

from auth.rbac import require_roles
from db import get_connection
from schemas.nomenclature import (
    ImportMode,
    NomenclatureCreate,
    NomenclatureImportCommitResponse,
    NomenclatureImportCommitRow,
    NomenclatureImportPreviewResponse,
    NomenclatureImportPreviewRow,
    NomenclatureRead,
    NomenclatureUpdate,
)
from schemas.nomenclature_route_chain import NomenclatureRouteChainResponse


router = APIRouter(prefix="/nomenclature", tags=["nomenclature"])

NOMENCLATURE_READ_ROLES = ("planner", "maintenance", "viewer")
NOMENCLATURE_WRITE_ROLES = ("planner",)

SELECT_COLUMNS = """
    nomenclature_id,
    nomenclature_code,
    nomenclature_name,
    unit_of_measure,
    is_active
"""


def has_item_type_column(cursor: RealDictCursor) -> bool:
    cursor.execute(
        """
        SELECT 1
        FROM information_schema.columns
        WHERE table_schema = 'public'
          AND table_name = 'nomenclature'
          AND column_name = 'item_type'
        LIMIT 1;
        """
    )
    return cursor.fetchone() is not None


def select_columns_with_item_type(item_type_column_exists: bool) -> str:
    if item_type_column_exists:
        return f"{SELECT_COLUMNS}, item_type"
    return f"{SELECT_COLUMNS}, 'manufactured'::text AS item_type"


MAX_ROUTE_CHAIN_DEPTH = 10
ROUTE_CHAIN_ROOT_NO_ACTIVE_ROUTE_WARNING = "Активный маршрут получения не найден."
ROUTE_CHAIN_MULTIPLE_ACTIVE_ROUTES_WARNING = (
    "Для позиции найдено несколько активных маршрутов. Использован первый."
)
ROUTE_CHAIN_MAX_DEPTH_WARNING = "Превышена максимальная глубина раскрытия маршрута."


def select_item_type_expression(item_type_column_exists: bool, table_alias: str) -> str:
    if item_type_column_exists:
        return f"{table_alias}.item_type"
    return "'manufactured'::text"


def append_warning(warnings: list[str], warning: str) -> None:
    if warning not in warnings:
        warnings.append(warning)


def format_nomenclature_label(code: str | None, name: str | None) -> str:
    code_value = (code or "").strip()
    name_value = (name or "").strip()

    if code_value and name_value:
        return f"{code_value} {name_value}"
    if code_value:
        return code_value
    if name_value:
        return name_value
    return "без кода"


def build_missing_component_route_warning(code: str | None, name: str | None) -> str:
    component_label = format_nomenclature_label(code, name)
    return f"Для производимого компонента {component_label} не найден активный маршрут."


def build_multiple_active_routes_warning(code: str | None, name: str | None) -> str:
    component_label = format_nomenclature_label(code, name)
    return f"Для позиции {component_label} найдено несколько активных маршрутов. Использован первый."


def build_cycle_warning(path_codes: list[str], repeated_code: str) -> str:
    cycle_path = " -> ".join(path_codes + [repeated_code])
    return f"Обнаружена циклическая зависимость маршрутов: {cycle_path}."


def build_route_chain_nomenclature_node(nomenclature_row: dict) -> dict:
    item_type = (nomenclature_row.get("item_type") or "manufactured").strip().lower()
    if item_type not in {"manufactured", "purchased"}:
        item_type = "manufactured"

    return {
        "nomenclature_id": nomenclature_row["nomenclature_id"],
        "nomenclature_code": nomenclature_row["nomenclature_code"],
        "nomenclature_name": nomenclature_row["nomenclature_name"],
        "unit_of_measure": nomenclature_row["unit_of_measure"],
        "item_type": item_type,
        "route": None,
    }


def fetch_nomenclature_for_route_chain(
    cursor: RealDictCursor,
    nomenclature_id: int,
    item_type_column_exists: bool,
) -> dict | None:
    item_type_expression = select_item_type_expression(item_type_column_exists, "n")
    cursor.execute(
        f"""
        SELECT
            n.nomenclature_id,
            n.nomenclature_code,
            n.nomenclature_name,
            n.unit_of_measure,
            {item_type_expression} AS item_type
        FROM nomenclature AS n
        WHERE n.nomenclature_id = %s;
        """,
        (nomenclature_id,),
    )
    return cursor.fetchone()


def fetch_active_routes_for_nomenclature(cursor: RealDictCursor, nomenclature_id: int) -> list[dict]:
    cursor.execute(
        """
        SELECT
            route_id,
            route_code,
            route_name,
            is_active
        FROM routes
        WHERE result_nomenclature_id = %s
          AND is_active = TRUE
        ORDER BY route_id;
        """,
        (nomenclature_id,),
    )
    return cursor.fetchall()


def fetch_route_steps_for_chain(cursor: RealDictCursor, route_id: int) -> list[dict]:
    cursor.execute(
        """
        SELECT
            rs.route_step_id,
            rs.step_no,
            rs.process_id,
            p.process_code,
            p.process_name,
            rs.output_nomenclature_id,
            n.nomenclature_code AS output_nomenclature_code,
            n.nomenclature_name AS output_nomenclature_name,
            rs.output_qty,
            rs.post_process_wait_hours
        FROM route_steps AS rs
        INNER JOIN processes AS p ON p.process_id = rs.process_id
        INNER JOIN nomenclature AS n ON n.nomenclature_id = rs.output_nomenclature_id
        WHERE rs.route_id = %s
        ORDER BY rs.step_no, rs.route_step_id;
        """,
        (route_id,),
    )
    return cursor.fetchall()


def fetch_route_step_inputs_for_chain(
    cursor: RealDictCursor,
    route_step_ids: list[int],
    item_type_column_exists: bool,
) -> dict[int, list[dict]]:
    if not route_step_ids:
        return {}

    item_type_expression = select_item_type_expression(item_type_column_exists, "n")
    cursor.execute(
        f"""
        SELECT
            rsi.route_step_id,
            rsi.step_input_id,
            rsi.input_nomenclature_id,
            rsi.input_qty,
            n.nomenclature_code AS input_nomenclature_code,
            n.nomenclature_name AS input_nomenclature_name,
            n.unit_of_measure,
            {item_type_expression} AS input_item_type
        FROM route_step_inputs AS rsi
        INNER JOIN nomenclature AS n ON n.nomenclature_id = rsi.input_nomenclature_id
        WHERE rsi.route_step_id = ANY(%s)
        ORDER BY rsi.route_step_id, rsi.step_input_id;
        """,
        (route_step_ids,),
    )
    rows = cursor.fetchall()

    grouped_rows: dict[int, list[dict]] = {}
    for row in rows:
        grouped_rows.setdefault(row["route_step_id"], []).append(row)

    return grouped_rows


def fetch_route_step_equipment_for_chain(
    cursor: RealDictCursor, route_step_ids: list[int]
) -> dict[int, list[dict]]:
    if not route_step_ids:
        return {}

    cursor.execute(
        """
        SELECT
            rse.route_step_id,
            rse.step_equipment_id,
            rse.machine_id,
            m.machine_code,
            m.machine_name,
            rse.equipment_role,
            rse.priority,
            rse.nominal_rate,
            rse.rate_uom,
            rse.min_batch_qty
        FROM route_step_equipment AS rse
        INNER JOIN machines AS m ON m.machine_id = rse.machine_id
        WHERE rse.route_step_id = ANY(%s)
        ORDER BY rse.route_step_id, rse.priority, rse.step_equipment_id;
        """,
        (route_step_ids,),
    )
    rows = cursor.fetchall()

    grouped_rows: dict[int, list[dict]] = {}
    for row in rows:
        grouped_rows.setdefault(row["route_step_id"], []).append(row)

    return grouped_rows


def build_nomenclature_route_chain(
    cursor: RealDictCursor,
    nomenclature_row: dict,
    item_type_column_exists: bool,
    warnings: list[str],
    path_nomenclature_ids: list[int],
    path_nomenclature_codes: list[str],
    depth: int,
    is_root: bool,
) -> dict | None:
    node = build_route_chain_nomenclature_node(nomenclature_row)

    active_routes = fetch_active_routes_for_nomenclature(cursor, node["nomenclature_id"])
    if not active_routes:
        if is_root:
            append_warning(warnings, ROUTE_CHAIN_ROOT_NO_ACTIVE_ROUTE_WARNING)
            return node
        else:
            append_warning(
                warnings,
                build_missing_component_route_warning(
                    node["nomenclature_code"],
                    node["nomenclature_name"],
                ),
            )
            return None

    if len(active_routes) > 1:
        if is_root:
            append_warning(warnings, ROUTE_CHAIN_MULTIPLE_ACTIVE_ROUTES_WARNING)
        else:
            append_warning(
                warnings,
                build_multiple_active_routes_warning(
                    node["nomenclature_code"],
                    node["nomenclature_name"],
                ),
            )

    selected_route = active_routes[0]
    route_steps = fetch_route_steps_for_chain(cursor, selected_route["route_id"])
    route_step_ids = [step["route_step_id"] for step in route_steps]

    inputs_by_step_id = fetch_route_step_inputs_for_chain(
        cursor=cursor,
        route_step_ids=route_step_ids,
        item_type_column_exists=item_type_column_exists,
    )
    equipment_by_step_id = fetch_route_step_equipment_for_chain(cursor, route_step_ids)

    steps_payload: list[dict] = []
    for step_row in route_steps:
        step_inputs_payload: list[dict] = []
        for input_row in inputs_by_step_id.get(step_row["route_step_id"], []):
            input_item_type = (input_row.get("input_item_type") or "manufactured").strip().lower()
            if input_item_type not in {"manufactured", "purchased"}:
                input_item_type = "manufactured"

            child_chain = None
            if input_item_type == "manufactured":
                input_nomenclature_id = int(input_row["input_nomenclature_id"])
                input_nomenclature_code = (
                    (input_row.get("input_nomenclature_code") or "").strip()
                    or f"ID:{input_nomenclature_id}"
                )

                if input_nomenclature_id in path_nomenclature_ids:
                    append_warning(
                        warnings,
                        build_cycle_warning(path_nomenclature_codes, input_nomenclature_code),
                    )
                elif depth >= MAX_ROUTE_CHAIN_DEPTH:
                    append_warning(warnings, ROUTE_CHAIN_MAX_DEPTH_WARNING)
                else:
                    child_chain = build_nomenclature_route_chain(
                        cursor=cursor,
                        nomenclature_row={
                            "nomenclature_id": input_nomenclature_id,
                            "nomenclature_code": input_row["input_nomenclature_code"],
                            "nomenclature_name": input_row["input_nomenclature_name"],
                            "unit_of_measure": input_row["unit_of_measure"],
                            "item_type": input_item_type,
                        },
                        item_type_column_exists=item_type_column_exists,
                        warnings=warnings,
                        path_nomenclature_ids=path_nomenclature_ids + [input_nomenclature_id],
                        path_nomenclature_codes=path_nomenclature_codes + [input_nomenclature_code],
                        depth=depth + 1,
                        is_root=False,
                    )

            step_inputs_payload.append(
                {
                    "step_input_id": input_row["step_input_id"],
                    "input_nomenclature_id": input_row["input_nomenclature_id"],
                    "input_nomenclature_code": input_row["input_nomenclature_code"],
                    "input_nomenclature_name": input_row["input_nomenclature_name"],
                    "input_item_type": input_item_type,
                    "input_qty": input_row["input_qty"],
                    "unit_of_measure": input_row["unit_of_measure"],
                    "child_chain": child_chain,
                }
            )

        step_equipment_payload = [
            {
                "step_equipment_id": equipment_row["step_equipment_id"],
                "machine_id": equipment_row["machine_id"],
                "machine_code": equipment_row["machine_code"],
                "machine_name": equipment_row["machine_name"],
                "equipment_role": equipment_row["equipment_role"],
                "priority": equipment_row["priority"],
                "nominal_rate": equipment_row["nominal_rate"],
                "rate_uom": equipment_row["rate_uom"],
                "min_batch_qty": equipment_row["min_batch_qty"],
            }
            for equipment_row in equipment_by_step_id.get(step_row["route_step_id"], [])
        ]

        steps_payload.append(
            {
                "route_step_id": step_row["route_step_id"],
                "step_no": step_row["step_no"],
                "process_id": step_row["process_id"],
                "process_code": step_row["process_code"],
                "process_name": step_row["process_name"],
                "output_nomenclature_id": step_row["output_nomenclature_id"],
                "output_nomenclature_code": step_row["output_nomenclature_code"],
                "output_nomenclature_name": step_row["output_nomenclature_name"],
                "output_qty": step_row["output_qty"],
                "post_process_wait_hours": step_row["post_process_wait_hours"],
                "inputs": step_inputs_payload,
                "equipment": step_equipment_payload,
            }
        )

    node["route"] = {
        "route_id": selected_route["route_id"],
        "route_code": selected_route["route_code"],
        "route_name": selected_route["route_name"],
        "is_active": selected_route["is_active"],
        "steps": steps_payload,
    }
    return node

IMPORT_MODE_ADD_ONLY: ImportMode = "add_only"
IMPORT_MODE_UPSERT: ImportMode = "upsert"
SUPPORTED_IMPORT_MODES: set[ImportMode] = {IMPORT_MODE_ADD_ONLY, IMPORT_MODE_UPSERT}

HEADER_FIELD_CODE = "nomenclature_code"
HEADER_FIELD_NAME = "nomenclature_name"
HEADER_FIELD_UNIT = "unit_of_measure"
HEADER_FIELD_ITEM_TYPE = "item_type"
HEADER_FIELD_ACTIVE = "is_active"

REQUIRED_IMPORT_HEADERS = {
    HEADER_FIELD_CODE,
    HEADER_FIELD_NAME,
    HEADER_FIELD_UNIT,
}

HEADER_ALIASES: dict[str, set[str]] = {
    HEADER_FIELD_CODE: {"РєРѕРґ", "nomenclaturecode"},
    HEADER_FIELD_NAME: {"РЅР°РёРјРµРЅРѕРІР°РЅРёРµ", "nomenclaturename"},
    HEADER_FIELD_UNIT: {"РµРґРёРЅРёС†Р°РёР·РјРµСЂРµРЅРёСЏ", "unitofmeasure"},
    HEADER_FIELD_ITEM_TYPE: {"типноменклатуры", "itemtype", "тип", "nomenclaturetype"},
    HEADER_FIELD_ACTIVE: {"Р°РєС‚РёРІРЅРѕСЃС‚СЊ", "isactive"},
}

TRUE_ACTIVE_VALUES = {
    "РґР°",
    "true",
    "1",
    "Р°РєС‚РёРІРЅР°",
    "Р°РєС‚РёРІРЅС‹Р№",
    "Р°РєС‚РёРІРЅРѕ",
    "yes",
    "y",
    "on",
}

FALSE_ACTIVE_VALUES = {
    "РЅРµС‚",
    "false",
    "0",
    "РЅРµР°РєС‚РёРІРЅР°",
    "РЅРµР°РєС‚РёРІРЅС‹Р№",
    "РЅРµР°РєС‚РёРІРЅРѕ",
    "no",
    "n",
    "off",
}


def normalize_header_name(value: object) -> str:
    if value is None:
        return ""

    normalized_value = str(value).strip().lower().replace("С‘", "Рµ")
    return re.sub(r"[\s_\-./\\]+", "", normalized_value)


def resolve_import_header(value: object) -> str | None:
    normalized_value = normalize_header_name(value)

    if not normalized_value:
        return None

    for field_name, aliases in HEADER_ALIASES.items():
        if normalized_value in aliases:
            return field_name

    return None


def normalize_import_mode(import_mode: str | None) -> ImportMode:
    normalized_mode = (import_mode or IMPORT_MODE_UPSERT).strip().lower()

    if normalized_mode not in SUPPORTED_IMPORT_MODES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="РќРµРґРѕРїСѓСЃС‚РёРјС‹Р№ СЂРµР¶РёРј РёРјРїРѕСЂС‚Р°. РСЃРїРѕР»СЊР·СѓР№С‚Рµ add_only РёР»Рё upsert.",
        )

    return normalized_mode  # type: ignore[return-value]


def normalize_bool_value(value: object) -> tuple[bool | None, str | None]:
    if value is None:
        return True, None

    if isinstance(value, bool):
        return value, None

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if value == 1:
            return True, None
        if value == 0:
            return False, None

    normalized_value = str(value).strip().lower()

    if not normalized_value:
        return True, None

    if normalized_value in TRUE_ACTIVE_VALUES:
        return True, None

    if normalized_value in FALSE_ACTIVE_VALUES:
        return False, None

    return None, "РќРµРґРѕРїСѓСЃС‚РёРјРѕРµ Р·РЅР°С‡РµРЅРёРµ Р°РєС‚РёРІРЅРѕСЃС‚Рё"


def normalize_unit_of_measure(value: object) -> tuple[str | None, str | None]:
    if value is None:
        return None, "Недопустимая единица измерения"

    raw_value = str(value).strip()
    if not raw_value:
        return None, "Недопустимая единица измерения"

    lowered_value = raw_value.lower()
    cyrillic_value = lowered_value.translate(str.maketrans({"m": "м", "p": "п"}))
    compact_value = re.sub(r"\s+", "", cyrillic_value)
    compact_without_dots = compact_value.replace(".", "")

    if compact_value in {"м²", "м2", "м^2", "m2"}:
        return "м²", None

    if compact_without_dots in {"мп", "мпог", "мпогон", "мпогонный"}:
        return "м.п.", None

    if compact_without_dots.startswith("м") and "пог" in compact_without_dots:
        return "м.п.", None

    if compact_without_dots in {"шт", "штука", "штуки", "pcs", "pc"}:
        return "шт", None

    if compact_without_dots in {"кг", "kg"}:
        return "кг", None

    if compact_without_dots in {"л", "литр", "литры", "l"}:
        return "л", None

    return None, "Недопустимая единица измерения"


def normalize_item_type(value: object) -> tuple[str, str | None]:
    if value is None:
        return "manufactured", None

    raw_value = str(value).strip().lower()
    if not raw_value:
        return "manufactured", None

    normalized = raw_value.replace("ё", "е")
    mapping = {
        "производимая": "manufactured",
        "manufactured": "manufactured",
        "prod": "manufactured",
        "закупаемая": "purchased",
        "purchased": "purchased",
        "buy": "purchased",
    }
    if normalized in mapping:
        return mapping[normalized], None

    return "manufactured", "Недопустимый тип номенклатуры"


def normalize_code(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_name(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def validate_import_file(file: UploadFile, file_bytes: bytes) -> None:
    file_name = file.filename or ""
    if not file_name.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="РџРѕРґРґРµСЂР¶РёРІР°РµС‚СЃСЏ С‚РѕР»СЊРєРѕ С„РѕСЂРјР°С‚ .xlsx.",
        )

    if len(file_bytes) == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Р¤Р°Р№Р» РїСѓСЃС‚РѕР№.",
        )


def read_import_rows(file_bytes: bytes) -> list[dict]:
    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="РќРµ СѓРґР°Р»РѕСЃСЊ РїСЂРѕС‡РёС‚Р°С‚СЊ Excel-С„Р°Р№Р». РџСЂРѕРІРµСЂСЊС‚Рµ С„РѕСЂРјР°С‚ .xlsx.",
        ) from exc

    try:
        sheet = workbook.active
        max_column = sheet.max_column or 0
        max_row = sheet.max_row or 0

        if max_column == 0 or max_row == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Р¤Р°Р№Р» РїСѓСЃС‚РѕР№.",
            )

        header_indexes: dict[str, int] = {}
        for column_index in range(1, max_column + 1):
            header_name = resolve_import_header(sheet.cell(row=1, column=column_index).value)
            if header_name and header_name not in header_indexes:
                header_indexes[header_name] = column_index

        missing_headers = [
            field_name
            for field_name in REQUIRED_IMPORT_HEADERS
            if field_name not in header_indexes
        ]
        if missing_headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "РќРµ РЅР°Р№РґРµРЅС‹ РѕР±СЏР·Р°С‚РµР»СЊРЅС‹Рµ РєРѕР»РѕРЅРєРё С€Р°Р±Р»РѕРЅР°: РљРѕРґ, РќР°РёРјРµРЅРѕРІР°РЅРёРµ, Р•РґРёРЅРёС†Р° РёР·РјРµСЂРµРЅРёСЏ."
                ),
            )

        rows: list[dict] = []
        has_active_column = HEADER_FIELD_ACTIVE in header_indexes
        has_item_type_column = HEADER_FIELD_ITEM_TYPE in header_indexes

        for row_index in range(2, max_row + 1):
            raw_code = sheet.cell(row=row_index, column=header_indexes[HEADER_FIELD_CODE]).value
            raw_name = sheet.cell(row=row_index, column=header_indexes[HEADER_FIELD_NAME]).value
            raw_unit = sheet.cell(row=row_index, column=header_indexes[HEADER_FIELD_UNIT]).value
            raw_item_type = (
                sheet.cell(row=row_index, column=header_indexes[HEADER_FIELD_ITEM_TYPE]).value
                if has_item_type_column
                else None
            )
            raw_active = (
                sheet.cell(row=row_index, column=header_indexes[HEADER_FIELD_ACTIVE]).value
                if has_active_column
                else None
            )

            is_empty_row = (
                normalize_code(raw_code) == ""
                and normalize_name(raw_name) == ""
                and normalize_code(raw_unit) == ""
                and (raw_item_type is None or str(raw_item_type).strip() == "")
                and (raw_active is None or str(raw_active).strip() == "")
            )
            if is_empty_row:
                continue

            normalized_code = normalize_code(raw_code)
            normalized_name = normalize_name(raw_name)
            normalized_unit, unit_error = normalize_unit_of_measure(raw_unit)
            normalized_item_type, item_type_error = normalize_item_type(raw_item_type)
            normalized_active, active_error = normalize_bool_value(raw_active)

            row_errors: list[str] = []

            if not normalized_code:
                row_errors.append("РџСѓСЃС‚РѕР№ РєРѕРґ")

            if not normalized_name:
                row_errors.append("РџСѓСЃС‚РѕРµ РЅР°РёРјРµРЅРѕРІР°РЅРёРµ")

            if unit_error:
                row_errors.append(unit_error)

            if item_type_error:
                row_errors.append(item_type_error)

            if active_error:
                row_errors.append(active_error)

            code_key = normalized_code.upper() if normalized_code else None

            rows.append(
                {
                    "row_no": row_index,
                    "nomenclature_code": normalized_code or None,
                    "nomenclature_name": normalized_name or None,
                    "unit_of_measure": normalized_unit,
                    "item_type": normalized_item_type,
                    "is_active": normalized_active,
                    "errors": row_errors,
                    "code_key": code_key,
                    "unit_normalized_from": (
                        str(raw_unit).strip()
                        if raw_unit is not None and normalized_unit is not None and str(raw_unit).strip() != normalized_unit
                        else None
                    ),
                }
            )

        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Р¤Р°Р№Р» РїСѓСЃС‚РѕР№.",
            )

        duplicate_counts: dict[str, int] = {}
        for row in rows:
            code_key = row["code_key"]
            if code_key:
                duplicate_counts[code_key] = duplicate_counts.get(code_key, 0) + 1

        for row in rows:
            code_key = row["code_key"]
            if code_key and duplicate_counts.get(code_key, 0) > 1:
                row["errors"].append("Р”СѓР±Р»РёРєР°С‚ РєРѕРґР° РІ С„Р°Р№Р»Рµ")

        return rows
    finally:
        workbook.close()


def fetch_existing_nomenclature_codes(cursor: RealDictCursor) -> dict[str, int]:
    cursor.execute(
        """
        SELECT nomenclature_id, nomenclature_code
        FROM nomenclature;
        """
    )
    rows = cursor.fetchall()
    return {
        str(row["nomenclature_code"]).strip().upper(): int(row["nomenclature_id"])
        for row in rows
    }


def build_import_preview(
    import_mode: ImportMode,
    parsed_rows: list[dict],
    existing_codes_map: dict[str, int],
) -> NomenclatureImportPreviewResponse:
    preview_rows: list[NomenclatureImportPreviewRow] = []
    new_rows = 0
    update_rows = 0
    conflict_rows = 0
    error_rows = 0
    valid_rows = 0

    for row in parsed_rows:
        row_errors: list[str] = list(row["errors"])
        code_key = row["code_key"]
        code_exists = bool(code_key and code_key in existing_codes_map)

        status_value: str
        can_import = False

        if row_errors:
            status_value = "error"
        elif code_exists and import_mode == IMPORT_MODE_ADD_ONLY:
            status_value = "conflict"
            row_errors.append("РљРѕРґ СѓР¶Рµ СЃСѓС‰РµСЃС‚РІСѓРµС‚")
        elif code_exists:
            status_value = "update"
            can_import = True
        else:
            status_value = "new"
            can_import = True

        if status_value == "new":
            new_rows += 1
        elif status_value == "update":
            update_rows += 1
        elif status_value == "conflict":
            conflict_rows += 1
        else:
            error_rows += 1

        if can_import:
            valid_rows += 1

        preview_rows.append(
            NomenclatureImportPreviewRow(
                row_no=row["row_no"],
                nomenclature_code=row["nomenclature_code"],
                nomenclature_name=row["nomenclature_name"],
                unit_of_measure=row["unit_of_measure"],
                item_type=row["item_type"],
                is_active=row["is_active"],
                status=status_value,  # type: ignore[arg-type]
                can_import=can_import,
                messages=row_errors,
                unit_normalized_from=row["unit_normalized_from"],
            )
        )

    return NomenclatureImportPreviewResponse(
        import_mode=import_mode,
        total_rows=len(parsed_rows),
        valid_rows=valid_rows,
        new_rows=new_rows,
        update_rows=update_rows,
        conflict_rows=conflict_rows,
        error_rows=error_rows,
        rows=preview_rows,
    )


def create_template_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "РќРѕРјРµРЅРєР»Р°С‚СѓСЂР°"
    sheet.append(["Код", "Наименование", "Единица измерения", "Тип номенклатуры", "Активность"])
    sheet.append(["NM-001", "Полотно ламинированное белое", "м²", "Производимая", "Да"])

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output.read()


def upsert_nomenclature_row(cursor: RealDictCursor, row: NomenclatureImportPreviewRow) -> bool:
    cursor.execute(
        """
        INSERT INTO nomenclature (
            nomenclature_code,
            nomenclature_name,
            unit_of_measure,
            item_type,
            is_active
        )
        VALUES (%s, %s, %s, %s, %s)
        ON CONFLICT (nomenclature_code)
        DO UPDATE SET
            nomenclature_name = EXCLUDED.nomenclature_name,
            unit_of_measure = EXCLUDED.unit_of_measure,
            item_type = EXCLUDED.item_type,
            is_active = EXCLUDED.is_active,
            updated_at = NOW()
        RETURNING (xmax = 0) AS inserted;
        """,
        (
            row.nomenclature_code,
            row.nomenclature_name,
            row.unit_of_measure,
            row.item_type,
            row.is_active,
        ),
    )
    result = cursor.fetchone()
    return bool(result and result["inserted"])


@router.get(
    "",
    response_model=List[NomenclatureRead],
    dependencies=[Depends(require_roles(*NOMENCLATURE_READ_ROLES))],
)
def list_nomenclature():
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            item_type_column_exists = has_item_type_column(cursor)
            cursor.execute(
                f"""
                SELECT {select_columns_with_item_type(item_type_column_exists)}
                FROM nomenclature
                ORDER BY nomenclature_code;
                """
            )
            rows = cursor.fetchall()

        return rows
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="РќРµ СѓРґР°Р»РѕСЃСЊ РїРѕР»СѓС‡РёС‚СЊ СЃРїРёСЃРѕРє РЅРѕРјРµРЅРєР»Р°С‚СѓСЂС‹.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.get(
    "/import/template",
    dependencies=[Depends(require_roles(*NOMENCLATURE_READ_ROLES))],
)
def download_nomenclature_import_template():
    template_content = create_template_workbook()
    return StreamingResponse(
        BytesIO(template_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="nomenclature_import_template.xlsx"'
        },
    )


@router.post(
    "/import/preview",
    response_model=NomenclatureImportPreviewResponse,
    dependencies=[Depends(require_roles(*NOMENCLATURE_WRITE_ROLES))],
)
async def preview_nomenclature_import(
    file: UploadFile = File(...),
    import_mode: str = Form(IMPORT_MODE_UPSERT),
):
    normalized_mode = normalize_import_mode(import_mode)
    file_bytes = await file.read()
    validate_import_file(file, file_bytes)
    parsed_rows = read_import_rows(file_bytes)

    connection = None
    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            existing_codes_map = fetch_existing_nomenclature_codes(cursor)

        return build_import_preview(
            import_mode=normalized_mode,
            parsed_rows=parsed_rows,
            existing_codes_map=existing_codes_map,
        )
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="РќРµ СѓРґР°Р»РѕСЃСЊ РїРѕРґРіРѕС‚РѕРІРёС‚СЊ РїСЂРµРґРїСЂРѕСЃРјРѕС‚СЂ РёРјРїРѕСЂС‚Р°.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.post(
    "/import/commit",
    response_model=NomenclatureImportCommitResponse,
    dependencies=[Depends(require_roles(*NOMENCLATURE_WRITE_ROLES))],
)
async def commit_nomenclature_import(
    file: UploadFile = File(...),
    import_mode: str = Form(IMPORT_MODE_UPSERT),
):
    normalized_mode = normalize_import_mode(import_mode)
    file_bytes = await file.read()
    validate_import_file(file, file_bytes)
    parsed_rows = read_import_rows(file_bytes)

    connection = None
    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            existing_codes_map = fetch_existing_nomenclature_codes(cursor)
            preview_result = build_import_preview(
                import_mode=normalized_mode,
                parsed_rows=parsed_rows,
                existing_codes_map=existing_codes_map,
            )

            created_count = 0
            updated_count = 0
            skipped_count = preview_result.conflict_rows + preview_result.error_rows
            error_count = preview_result.error_rows
            conflict_count = preview_result.conflict_rows
            commit_rows: list[NomenclatureImportCommitRow] = []

            for preview_row in preview_result.rows:
                if preview_row.status == "error":
                    commit_rows.append(
                        NomenclatureImportCommitRow(
                            row_no=preview_row.row_no,
                            nomenclature_code=preview_row.nomenclature_code,
                            status="error",
                            message=preview_row.messages[0] if preview_row.messages else "РћС€РёР±РєР° РІР°Р»РёРґР°С†РёРё",
                        )
                    )
                    continue

                if preview_row.status == "conflict":
                    commit_rows.append(
                        NomenclatureImportCommitRow(
                            row_no=preview_row.row_no,
                            nomenclature_code=preview_row.nomenclature_code,
                            status="skipped",
                            message=preview_row.messages[0] if preview_row.messages else "РљРѕРґ СѓР¶Рµ СЃСѓС‰РµСЃС‚РІСѓРµС‚",
                        )
                    )
                    continue

                if normalized_mode == IMPORT_MODE_ADD_ONLY:
                    cursor.execute(
                        """
                        INSERT INTO nomenclature (
                            nomenclature_code,
                            nomenclature_name,
                            unit_of_measure,
                            item_type,
                            is_active
                        )
                        VALUES (%s, %s, %s, %s, %s)
                        ON CONFLICT (nomenclature_code)
                        DO NOTHING
                        RETURNING nomenclature_id;
                        """,
                        (
                            preview_row.nomenclature_code,
                            preview_row.nomenclature_name,
                            preview_row.unit_of_measure,
                            preview_row.item_type,
                            preview_row.is_active,
                        ),
                    )
                    inserted_row = cursor.fetchone()
                    if inserted_row is None:
                        skipped_count += 1
                        conflict_count += 1
                        commit_rows.append(
                            NomenclatureImportCommitRow(
                                row_no=preview_row.row_no,
                                nomenclature_code=preview_row.nomenclature_code,
                                status="skipped",
                                message="РљРѕРґ СѓР¶Рµ СЃСѓС‰РµСЃС‚РІСѓРµС‚",
                            )
                        )
                    else:
                        created_count += 1
                        commit_rows.append(
                            NomenclatureImportCommitRow(
                                row_no=preview_row.row_no,
                                nomenclature_code=preview_row.nomenclature_code,
                                status="created",
                                message="РЎРѕР·РґР°РЅРѕ",
                            )
                        )
                    continue

                inserted = upsert_nomenclature_row(cursor, preview_row)
                if inserted:
                    created_count += 1
                    commit_rows.append(
                        NomenclatureImportCommitRow(
                            row_no=preview_row.row_no,
                            nomenclature_code=preview_row.nomenclature_code,
                            status="created",
                            message="РЎРѕР·РґР°РЅРѕ",
                        )
                    )
                else:
                    updated_count += 1
                    commit_rows.append(
                        NomenclatureImportCommitRow(
                            row_no=preview_row.row_no,
                            nomenclature_code=preview_row.nomenclature_code,
                            status="updated",
                            message="РћР±РЅРѕРІР»РµРЅРѕ",
                        )
                    )

        connection.commit()
        return NomenclatureImportCommitResponse(
            import_mode=normalized_mode,
            total_rows=preview_result.total_rows,
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=skipped_count,
            error_count=error_count,
            conflict_count=conflict_count,
            rows=commit_rows,
        )
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="РќРµ СѓРґР°Р»РѕСЃСЊ РІС‹РїРѕР»РЅРёС‚СЊ РёРјРїРѕСЂС‚ РЅРѕРјРµРЅРєР»Р°С‚СѓСЂС‹.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.get(
    "/{nomenclature_id}/route-chain",
    response_model=NomenclatureRouteChainResponse,
    dependencies=[Depends(require_roles(*NOMENCLATURE_READ_ROLES))],
)
def get_nomenclature_route_chain(nomenclature_id: int = Path(..., gt=0)):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            item_type_column_exists = has_item_type_column(cursor)
            nomenclature_row = fetch_nomenclature_for_route_chain(
                cursor=cursor,
                nomenclature_id=nomenclature_id,
                item_type_column_exists=item_type_column_exists,
            )

            if nomenclature_row is None:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Позиция номенклатуры не найдена.",
                )

            root_code = (
                (nomenclature_row.get("nomenclature_code") or "").strip()
                or f"ID:{nomenclature_row['nomenclature_id']}"
            )
            warnings: list[str] = []
            route_chain = build_nomenclature_route_chain(
                cursor=cursor,
                nomenclature_row=nomenclature_row,
                item_type_column_exists=item_type_column_exists,
                warnings=warnings,
                path_nomenclature_ids=[nomenclature_row["nomenclature_id"]],
                path_nomenclature_codes=[root_code],
                depth=1,
                is_root=True,
            )
            if route_chain is None:
                route_chain = build_route_chain_nomenclature_node(nomenclature_row)

        route_chain["warnings"] = warnings
        return route_chain
    except HTTPException:
        raise
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось получить полную цепочку маршрута номенклатуры.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.get(
    "/{nomenclature_id}",
    response_model=NomenclatureRead,
    dependencies=[Depends(require_roles(*NOMENCLATURE_READ_ROLES))],
)
def get_nomenclature(nomenclature_id: int = Path(..., gt=0)):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            item_type_column_exists = has_item_type_column(cursor)
            cursor.execute(
                f"""
                SELECT {select_columns_with_item_type(item_type_column_exists)}
                FROM nomenclature
                WHERE nomenclature_id = %s;
                """,
                (nomenclature_id,),
            )
            row = cursor.fetchone()

        if row is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="РџРѕР·РёС†РёСЏ РЅРѕРјРµРЅРєР»Р°С‚СѓСЂС‹ РЅРµ РЅР°Р№РґРµРЅР°.",
            )

        return row
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="РќРµ СѓРґР°Р»РѕСЃСЊ РїРѕР»СѓС‡РёС‚СЊ РїРѕР·РёС†РёСЋ РЅРѕРјРµРЅРєР»Р°С‚СѓСЂС‹.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.post(
    "",
    response_model=NomenclatureRead,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(require_roles(*NOMENCLATURE_WRITE_ROLES))],
)
def create_nomenclature(payload: NomenclatureCreate):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            item_type_column_exists = has_item_type_column(cursor)
            if item_type_column_exists:
                cursor.execute(
                    f"""
                    INSERT INTO nomenclature (
                        nomenclature_code,
                        nomenclature_name,
                        unit_of_measure,
                        item_type,
                        is_active
                    )
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING {select_columns_with_item_type(True)};
                    """,
                    (
                        payload.nomenclature_code,
                        payload.nomenclature_name,
                        payload.unit_of_measure,
                        payload.item_type,
                        payload.is_active,
                    ),
                )
            else:
                cursor.execute(
                    f"""
                    INSERT INTO nomenclature (
                        nomenclature_code,
                        nomenclature_name,
                        unit_of_measure,
                        is_active
                    )
                    VALUES (%s, %s, %s, %s)
                    RETURNING {select_columns_with_item_type(False)};
                    """,
                    (
                        payload.nomenclature_code,
                        payload.nomenclature_name,
                        payload.unit_of_measure,
                        payload.is_active,
                    ),
                )
            created_row = cursor.fetchone()

        connection.commit()
        return created_row
    except UniqueViolation as exc:
        if connection is not None:
            connection.rollback()

        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="РџРѕР·РёС†РёСЏ СЃ С‚Р°РєРёРј РєРѕРґРѕРј СѓР¶Рµ СЃСѓС‰РµСЃС‚РІСѓРµС‚.",
        ) from exc
    except CheckViolation as exc:
        if connection is not None:
            connection.rollback()

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Р•РґРёРЅРёС†Р° РёР·РјРµСЂРµРЅРёСЏ РјРѕР¶РµС‚ Р±С‹С‚СЊ С‚РѕР»СЊРєРѕ 'РјВІ' РёР»Рё 'Рј.Рї.'.",
        ) from exc
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="РќРµ СѓРґР°Р»РѕСЃСЊ СЃРѕР·РґР°С‚СЊ РїРѕР·РёС†РёСЋ РЅРѕРјРµРЅРєР»Р°С‚СѓСЂС‹.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.put(
    "/{nomenclature_id}",
    response_model=NomenclatureRead,
    dependencies=[Depends(require_roles(*NOMENCLATURE_WRITE_ROLES))],
)
def update_nomenclature(
    payload: NomenclatureUpdate,
    nomenclature_id: int = Path(..., gt=0),
):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            item_type_column_exists = has_item_type_column(cursor)
            if item_type_column_exists:
                cursor.execute(
                    f"""
                    UPDATE nomenclature
                    SET
                        nomenclature_code = %s,
                        nomenclature_name = %s,
                        unit_of_measure = %s,
                        item_type = %s,
                        is_active = %s,
                        updated_at = NOW()
                    WHERE nomenclature_id = %s
                    RETURNING {select_columns_with_item_type(True)};
                    """,
                    (
                        payload.nomenclature_code,
                        payload.nomenclature_name,
                        payload.unit_of_measure,
                        payload.item_type,
                        payload.is_active,
                        nomenclature_id,
                    ),
                )
            else:
                cursor.execute(
                    f"""
                    UPDATE nomenclature
                    SET
                        nomenclature_code = %s,
                        nomenclature_name = %s,
                        unit_of_measure = %s,
                        is_active = %s,
                        updated_at = NOW()
                    WHERE nomenclature_id = %s
                    RETURNING {select_columns_with_item_type(False)};
                    """,
                    (
                        payload.nomenclature_code,
                        payload.nomenclature_name,
                        payload.unit_of_measure,
                        payload.is_active,
                        nomenclature_id,
                    ),
                )
            updated_row = cursor.fetchone()

        if updated_row is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="РџРѕР·РёС†РёСЏ РЅРѕРјРµРЅРєР»Р°С‚СѓСЂС‹ РЅРµ РЅР°Р№РґРµРЅР°.",
            )

        connection.commit()
        return updated_row
    except UniqueViolation as exc:
        if connection is not None:
            connection.rollback()

        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="РџРѕР·РёС†РёСЏ СЃ С‚Р°РєРёРј РєРѕРґРѕРј СѓР¶Рµ СЃСѓС‰РµСЃС‚РІСѓРµС‚.",
        ) from exc
    except CheckViolation as exc:
        if connection is not None:
            connection.rollback()

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Р•РґРёРЅРёС†Р° РёР·РјРµСЂРµРЅРёСЏ РјРѕР¶РµС‚ Р±С‹С‚СЊ С‚РѕР»СЊРєРѕ 'РјВІ' РёР»Рё 'Рј.Рї.'.",
        ) from exc
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="РќРµ СѓРґР°Р»РѕСЃСЊ РёР·РјРµРЅРёС‚СЊ РїРѕР·РёС†РёСЋ РЅРѕРјРµРЅРєР»Р°С‚СѓСЂС‹.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.delete(
    "/{nomenclature_id}",
    response_model=NomenclatureRead,
    dependencies=[Depends(require_roles(*NOMENCLATURE_WRITE_ROLES))],
)
def deactivate_nomenclature(nomenclature_id: int = Path(..., gt=0)):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            item_type_column_exists = has_item_type_column(cursor)
            cursor.execute(
                f"""
                UPDATE nomenclature
                SET
                    is_active = FALSE,
                    updated_at = NOW()
                WHERE nomenclature_id = %s
                RETURNING {select_columns_with_item_type(item_type_column_exists)};
                """,
                (nomenclature_id,),
            )
            deactivated_row = cursor.fetchone()

        if deactivated_row is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="РџРѕР·РёС†РёСЏ РЅРѕРјРµРЅРєР»Р°С‚СѓСЂС‹ РЅРµ РЅР°Р№РґРµРЅР°.",
            )

        connection.commit()
        return deactivated_row
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="РќРµ СѓРґР°Р»РѕСЃСЊ РґРµР°РєС‚РёРІРёСЂРѕРІР°С‚СЊ РїРѕР·РёС†РёСЋ РЅРѕРјРµРЅРєР»Р°С‚СѓСЂС‹.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()







