import re
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from typing import Any
from urllib.parse import quote

import psycopg2
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from psycopg2.extras import RealDictCursor

from auth.rbac import require_roles
from db import get_connection
from routers.equipment_downtimes import get_local_now
from routers.production_week_plans import build_equipment_availability
from schemas.production_analytics import (
    EquipmentMonthlyAnalyticsResponse,
    EquipmentMonthlyAnalyticsSummary,
    EquipmentMonthlyDowntimeCategoryItem,
    EquipmentMonthlyDowntimeItem,
    EquipmentMonthlyLoadItem,
    MonthlyOutputAnalyticsItem,
    MonthlyOutputAnalyticsProblemItem,
    MonthlyOutputAnalyticsResponse,
    MonthlyOutputAnalyticsSummary,
    MonthlyOutputAnalyticsSummaryByUnit,
)


router = APIRouter(prefix="/production-analytics", tags=["production_analytics"])

ANALYTICS_READ_ROLES = ("planner", "master", "maintenance", "viewer")

DECIMAL_ZERO = Decimal("0")
MONTH_PATTERN = re.compile(r"^\d{4}-\d{2}$")
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

STATUS_NO_ACTUAL = "no_actual"
STATUS_IN_PROGRESS = "in_progress"
STATUS_COMPLETED = "completed"
STATUS_OVERPRODUCED = "overproduced"
STATUS_NO_PLAN = "no_plan"

STATUS_LABELS = {
    STATUS_NO_ACTUAL: "Нет факта",
    STATUS_IN_PROGRESS: "В работе",
    STATUS_COMPLETED: "Выполнено",
    STATUS_OVERPRODUCED: "Перевыпуск",
    STATUS_NO_PLAN: "Нет плана",
}

EQUIPMENT_STATUS_NO_DATA = "no_capacity_data"
EQUIPMENT_STATUS_OVERLOADED = "overloaded"
EQUIPMENT_STATUS_HIGH_LOAD = "high_load"
EQUIPMENT_STATUS_NORMAL = "normal"
EQUIPMENT_STATUS_NO_LOAD = "no_load"

EQUIPMENT_STATUS_LABELS = {
    EQUIPMENT_STATUS_NO_DATA: "Нет данных",
    EQUIPMENT_STATUS_OVERLOADED: "Перегруз",
    EQUIPMENT_STATUS_HIGH_LOAD: "Высокая загрузка",
    EQUIPMENT_STATUS_NORMAL: "Норма",
}


def parse_month_value(value: str) -> date:
    normalized_value = str(value or "").strip()
    if not MONTH_PATTERN.fullmatch(normalized_value):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Параметр month должен быть в формате YYYY-MM.",
        )

    try:
        year, month = normalized_value.split("-")
        return date(int(year), int(month), 1)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Параметр month должен быть в формате YYYY-MM.",
        ) from exc


def get_next_month(month_start: date) -> date:
    if month_start.month == 12:
        return date(month_start.year + 1, 1, 1)
    return date(month_start.year, month_start.month + 1, 1)


def get_month_period_bounds(month_start: date) -> tuple[date, date, datetime, datetime]:
    month_end = get_next_month(month_start)
    date_from = datetime.combine(month_start, time.min)
    date_to = datetime.combine(month_end, time.min)
    return month_start, month_end, date_from, date_to


def to_decimal(value: Any) -> Decimal:
    if value is None:
        return DECIMAL_ZERO
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def to_qty_float(value: Decimal) -> float:
    return round(float(value), 3)


def to_hours_float(value: Decimal | int | float | None) -> float:
    if value is None:
        return 0.0
    decimal_value = to_decimal(value)
    return round(float(decimal_value / Decimal("60")), 1)


def to_percent_float(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return round(float(value), 1)


def normalize_unit_of_measure(value: Any) -> str:
    unit = str(value or "").strip()
    return unit or "Без ед."


def build_empty_summary() -> MonthlyOutputAnalyticsSummary:
    return MonthlyOutputAnalyticsSummary(
        planned_qty_total=0.0,
        actual_qty_total=0.0,
        remaining_qty_total=0.0,
        completion_percent=0.0,
        underproduced_items_count=0,
        overproduced_items_count=0,
        no_actual_items_count=0,
        no_plan_items_count=0,
    )


def build_empty_equipment_summary() -> EquipmentMonthlyAnalyticsSummary:
    return EquipmentMonthlyAnalyticsSummary(
        equipment_in_plan_count=0,
        average_load_percent=0.0,
        overloaded_equipment_count=0,
        high_load_equipment_count=0,
        total_downtime_hours=0.0,
        planned_maintenance_hours=0.0,
        unplanned_downtime_hours=0.0,
        unplanned_share_percent=0.0,
    )


def build_item_status(planned_qty: Decimal, actual_qty: Decimal) -> tuple[str, str]:
    if planned_qty <= DECIMAL_ZERO and actual_qty > DECIMAL_ZERO:
        return STATUS_NO_PLAN, STATUS_LABELS[STATUS_NO_PLAN]
    if planned_qty > DECIMAL_ZERO and actual_qty <= DECIMAL_ZERO:
        return STATUS_NO_ACTUAL, STATUS_LABELS[STATUS_NO_ACTUAL]
    if planned_qty > DECIMAL_ZERO and actual_qty < planned_qty:
        return STATUS_IN_PROGRESS, STATUS_LABELS[STATUS_IN_PROGRESS]
    if planned_qty > DECIMAL_ZERO and actual_qty == planned_qty:
        return STATUS_COMPLETED, STATUS_LABELS[STATUS_COMPLETED]
    if planned_qty > DECIMAL_ZERO and actual_qty > planned_qty:
        return STATUS_OVERPRODUCED, STATUS_LABELS[STATUS_OVERPRODUCED]
    return STATUS_COMPLETED, STATUS_LABELS[STATUS_COMPLETED]


def build_equipment_status(
    available_minutes: int,
    planned_load_minutes: Decimal,
    has_capacity_gap: bool,
) -> tuple[str, str, str | None, Decimal | None]:
    if has_capacity_gap:
        return (
            EQUIPMENT_STATUS_NO_DATA,
            EQUIPMENT_STATUS_LABELS[EQUIPMENT_STATUS_NO_DATA],
            "Для части строк не задана производительность оборудования.",
            None,
        )

    if available_minutes <= 0:
        if planned_load_minutes > DECIMAL_ZERO:
            return (
                EQUIPMENT_STATUS_OVERLOADED,
                EQUIPMENT_STATUS_LABELS[EQUIPMENT_STATUS_OVERLOADED],
                "Оборудование перегружено",
                None,
            )
        return (
            EQUIPMENT_STATUS_NORMAL,
            EQUIPMENT_STATUS_LABELS[EQUIPMENT_STATUS_NORMAL],
            None,
            Decimal("0"),
        )

    load_percent = (planned_load_minutes / Decimal(available_minutes)) * Decimal("100")
    load_percent = load_percent.quantize(Decimal("0.1"))

    if load_percent > Decimal("100"):
        return (
            EQUIPMENT_STATUS_OVERLOADED,
            EQUIPMENT_STATUS_LABELS[EQUIPMENT_STATUS_OVERLOADED],
            "Оборудование перегружено",
            load_percent,
        )
    if load_percent >= Decimal("85"):
        return (
            EQUIPMENT_STATUS_HIGH_LOAD,
            EQUIPMENT_STATUS_LABELS[EQUIPMENT_STATUS_HIGH_LOAD],
            None,
            load_percent,
        )
    return (
        EQUIPMENT_STATUS_NORMAL,
        EQUIPMENT_STATUS_LABELS[EQUIPMENT_STATUS_NORMAL],
        None,
        load_percent,
    )


def fetch_week_lines_for_equipment_analytics(
    cursor: RealDictCursor,
    production_plan_week_id: int,
) -> list[dict[str, Any]]:
    cursor.execute(
        """
        SELECT
            pwl.production_week_line_id,
            pwl.route_step_equipment_id,
            rse.machine_id,
            m.machine_code,
            m.machine_name,
            pwl.planned_qty,
            rse.nominal_rate
        FROM production_week_lines AS pwl
        LEFT JOIN route_step_equipment AS rse ON rse.step_equipment_id = pwl.route_step_equipment_id
        LEFT JOIN machines AS m ON m.machine_id = rse.machine_id
        WHERE pwl.production_plan_week_id = %s
        ORDER BY pwl.sequence_no ASC, pwl.production_week_line_id ASC;
        """,
        (production_plan_week_id,),
    )
    return cursor.fetchall()


def fetch_planned_maintenance_minutes_for_period(
    cursor: RealDictCursor,
    date_from: datetime,
    date_to: datetime,
) -> Decimal:
    cursor.execute(
        """
        SELECT
            COALESCE(
                SUM(
                    GREATEST(
                        EXTRACT(
                            EPOCH FROM (
                                LEAST(em.ended_at, %s) - GREATEST(em.started_at, %s)
                            )
                        ) / 60,
                        0
                    )
                ),
                0
            ) AS maintenance_minutes
        FROM equipment_maintenance AS em
        WHERE em.started_at < %s
          AND em.ended_at > %s;
        """,
        (date_to, date_from, date_to, date_from),
    )
    row = cursor.fetchone()
    return to_decimal(row["maintenance_minutes"])


def get_field(item: Any, field_name: str, default: Any = None) -> Any:
    if isinstance(item, dict):
        return item.get(field_name, default)
    return getattr(item, field_name, default)


def format_print_datetime(value: datetime) -> str:
    return value.strftime("%d.%m.%Y %H:%M")


def format_ru_number(value: Any, minimum_fraction_digits: int = 0, maximum_fraction_digits: int = 1) -> str:
    if value is None:
        return "—"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "—"
    text = f"{number:,.{maximum_fraction_digits}f}"
    if minimum_fraction_digits == 0 and "." in text:
        text = text.rstrip("0").rstrip(".")
    integer_part, _, fractional_part = text.partition(".")
    integer_part = integer_part.replace(",", " ")
    return f"{integer_part},{fractional_part}" if fractional_part else integer_part


def format_qty_thousands_for_print(value: Any) -> str:
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        number = 0.0
    return f"{format_ru_number(number / 1000, 1, 1)} тыс."


def format_percent_for_print(value: Any) -> str:
    return f"{format_ru_number(value or 0, 0, 1)}%"


def format_qty_by_unit_for_print(summary_by_unit: list[Any], qty_field: str, fallback_value: Any) -> str:
    rows = list(summary_by_unit or [])
    if not rows:
        return format_qty_thousands_for_print(fallback_value)

    lines = []
    for row in rows:
        unit = str(get_field(row, "unit", "") or "").strip()
        suffix = f" {unit}" if unit else ""
        lines.append(f"{format_qty_thousands_for_print(get_field(row, qty_field, 0))}{suffix}")
    return "\n".join(lines)


def configure_print_sheet(worksheet, widths: list[float]) -> None:
    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.freeze_panes = "A7"
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(64 + index)].width = width


def write_report_header(worksheet, month: str, section_title: str, last_column: int) -> int:
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title_cell = worksheet.cell(row=1, column=1, value="Анализ выпуска")
    title_cell.font = Font(size=16, bold=True, color="1E293B")
    title_cell.alignment = Alignment(horizontal="center")
    worksheet.cell(row=2, column=1, value=f"Период: {month}").font = Font(size=10, color="475569")
    worksheet.cell(row=3, column=1, value=f"Дата формирования: {format_print_datetime(datetime.now())}").font = Font(size=10, color="475569")
    worksheet.cell(row=4, column=1, value=f"Раздел: {section_title}").font = Font(size=10, bold=True, color="475569")
    worksheet.row_dimensions[1].height = 24
    return 6


def write_section_title(worksheet, row_index: int, title: str, last_column: int) -> int:
    worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=last_column)
    cell = worksheet.cell(row=row_index, column=1, value=title)
    cell.font = Font(size=12, bold=True, color="0F172A")
    cell.fill = PatternFill("solid", fgColor="E8F1F5")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[row_index].height = 22
    return row_index + 1


def write_kpi_block(worksheet, row_index: int, items: list[tuple[str, Any]], columns_per_row: int = 4) -> int:
    header_fill = PatternFill("solid", fgColor="F1F5F9")
    thin_side = Side(style="thin", color="B7C6CE")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    label_font = Font(size=9, bold=True, color="475569")
    value_font = Font(size=11, bold=True, color="0F172A")

    for index, (label, value) in enumerate(items):
        column = (index % columns_per_row) * 2 + 1
        if index and index % columns_per_row == 0:
            row_index += 2
        label_cell = worksheet.cell(row=row_index, column=column, value=label)
        value_cell = worksheet.cell(row=row_index + 1, column=column, value=value)
        worksheet.merge_cells(start_row=row_index, start_column=column, end_row=row_index, end_column=column + 1)
        worksheet.merge_cells(start_row=row_index + 1, start_column=column, end_row=row_index + 1, end_column=column + 1)
        for cell in (label_cell, value_cell):
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        label_cell.font = label_font
        value_cell.font = value_font
    return row_index + 3


def write_table(
    worksheet,
    row_index: int,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    numeric_columns: set[int] | None = None,
    percent_columns: set[int] | None = None,
    wrap_columns: set[int] | None = None,
) -> int:
    numeric_columns = numeric_columns or set()
    percent_columns = percent_columns or set()
    wrap_columns = wrap_columns or set()
    last_column = len(headers)
    row_index = write_section_title(worksheet, row_index, title, last_column)

    header_fill = PatternFill("solid", fgColor="DDEBEE")
    thin_side = Side(style="thin", color="B7C6CE")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_font = Font(size=10, bold=True, color="0F172A")

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row_index, column=column_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if not rows:
        row_index += 1
        worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=last_column)
        cell = worksheet.cell(row=row_index, column=1, value="Данные за выбранный период отсутствуют.")
        cell.font = Font(size=10, color="475569")
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        return row_index + 3

    for row_values in rows:
        row_index += 1
        for column_index, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=column_index in wrap_columns)
            if column_index == 1:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=column_index in wrap_columns)
            if column_index in numeric_columns or column_index in percent_columns:
                cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=column_index in wrap_columns)
                cell.number_format = "0.0" if column_index in percent_columns else "#,##0.0"
    return row_index + 3


def get_completion_sort_value(item: Any) -> tuple[float, str, str]:
    completion_percent = get_field(item, "completion_percent")
    try:
        percent_value = float(completion_percent)
    except (TypeError, ValueError):
        percent_value = float("inf")
    if not (percent_value == percent_value):
        percent_value = float("inf")
    return (
        percent_value,
        str(get_field(item, "item_code", "") or ""),
        str(get_field(item, "item_name", "") or ""),
    )


def write_summary_table(
    worksheet,
    row_index: int,
    title: str,
    rows: list[list[Any]],
    last_column: int,
) -> int:
    value_start_column = max(2, last_column - 1)
    row_index = write_section_title(worksheet, row_index, title, last_column)

    header_fill = PatternFill("solid", fgColor="DDEBEE")
    total_fill = PatternFill("solid", fgColor="F8FAFC")
    thin_side = Side(style="thin", color="B7C6CE")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_font = Font(size=10, bold=True, color="0F172A")

    worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=value_start_column - 1)
    worksheet.merge_cells(start_row=row_index, start_column=value_start_column, end_row=row_index, end_column=last_column)
    indicator_header_cell = worksheet.cell(row=row_index, column=1, value="Показатель")
    value_header_cell = worksheet.cell(row=row_index, column=value_start_column, value="Значение")
    for cell in (indicator_header_cell, value_header_cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if not rows:
        row_index += 1
        worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=last_column)
        cell = worksheet.cell(row=row_index, column=1, value="Данные за выбранный период отсутствуют.")
        cell.font = Font(size=10, color="475569")
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        return row_index + 3

    for label, value in rows:
        row_index += 1
        worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=value_start_column - 1)
        worksheet.merge_cells(start_row=row_index, start_column=value_start_column, end_row=row_index, end_column=last_column)
        label_cell = worksheet.cell(row=row_index, column=1, value=label)
        value_cell = worksheet.cell(row=row_index, column=value_start_column, value=value)
        for cell in (label_cell, value_cell):
            cell.fill = total_fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        value_cell.alignment = Alignment(horizontal="right", vertical="top")
        value_cell.number_format = "#,##0.0"

    return row_index + 3


def create_production_analytics_print_workbook(
    month: str,
    output_data: MonthlyOutputAnalyticsResponse,
    equipment_data: EquipmentMonthlyAnalyticsResponse,
) -> bytes:
    workbook = Workbook()
    output_sheet = workbook.active
    output_sheet.title = "Выполнение плана"
    capacity_sheet = workbook.create_sheet("Обеспеченность мощностями")

    render_output_analytics_sheet(output_sheet, month, output_data)
    render_capacity_analytics_sheet(capacity_sheet, month, equipment_data)

    workbook_stream = BytesIO()
    workbook.save(workbook_stream)
    return workbook_stream.getvalue()


def render_output_analytics_sheet(worksheet, month: str, data: MonthlyOutputAnalyticsResponse) -> None:
    configure_print_sheet(worksheet, [5, 18, 42, 14, 14, 14, 14, 22])
    current_row = write_report_header(worksheet, month, "Выполнение плана", 8)
    summary = data.summary
    summary_by_unit = data.summary_by_unit or []
    current_row = write_kpi_block(
        worksheet,
        current_row,
        [
            ("План на месяц", format_qty_by_unit_for_print(summary_by_unit, "planned_qty_total", summary.planned_qty_total)),
            ("Выпущено", format_qty_by_unit_for_print(summary_by_unit, "actual_qty_total", summary.actual_qty_total)),
            ("Остаток к выпуску", format_qty_by_unit_for_print(summary_by_unit, "remaining_qty_total", summary.remaining_qty_total)),
            ("Выполнение, %", format_percent_for_print(summary.completion_percent)),
        ],
    )

    def output_rows(items: list[Any]) -> list[list[Any]]:
        return [
            [
                index,
                item.item_code,
                item.item_name,
                item.planned_qty,
                item.actual_qty,
                item.remaining_qty,
                item.completion_percent,
                item.status_label,
            ]
            for index, item in enumerate(items or [], start=1)
        ]

    headers = ["№", "Код", "Номенклатура", "План", "Факт", "Остаток", "Выполнение, %", "Статус"]
    current_row = write_table(
        worksheet,
        current_row,
        "Проблемные позиции",
        headers,
        output_rows(data.top_problem_items),
        numeric_columns={4, 5, 6},
        percent_columns={7},
        wrap_columns={3, 8},
    )
    current_row = write_table(
        worksheet,
        current_row,
        "План-факт по номенклатуре",
        headers,
        output_rows(sorted(data.items or [], key=get_completion_sort_value)),
        numeric_columns={4, 5, 6},
        percent_columns={7},
        wrap_columns={3, 8},
    )

    summary_rows = [
        ["Позиций с недовыпуском", summary.underproduced_items_count],
        ["Позиций без факта", summary.no_actual_items_count],
        ["Позиций с перевыпуском", summary.overproduced_items_count],
        ["Позиций без плана", summary.no_plan_items_count],
    ]
    write_summary_table(worksheet, current_row, "Сводка периода", summary_rows, last_column=8)


def render_capacity_analytics_sheet(worksheet, month: str, data: EquipmentMonthlyAnalyticsResponse) -> None:
    configure_print_sheet(worksheet, [5, 18, 30, 14, 18, 16, 18, 14, 22])
    current_row = write_report_header(worksheet, month, "Обеспеченность мощностями", 9)
    summary = data.summary
    current_row = write_kpi_block(
        worksheet,
        current_row,
        [
            ("Оборудование в плане", summary.equipment_in_plan_count),
            ("Средняя загрузка, %", format_percent_for_print(summary.average_load_percent)),
            ("Перегружено", summary.overloaded_equipment_count),
            ("Всего простоев, ч", format_ru_number(summary.total_downtime_hours, 0, 1)),
        ],
    )

    load_rows = [
        [
            index,
            item.equipment_code,
            item.equipment_name,
            item.available_hours,
            item.planned_load_hours if item.planned_load_hours is not None else "—",
            item.planned_maintenance_hours,
            item.remaining_hours,
            item.load_percent if item.load_percent is not None else "—",
            item.status_label,
        ]
        for index, item in enumerate(data.equipment_load or [], start=1)
    ]
    current_row = write_table(
        worksheet,
        current_row,
        "Расчётная загрузка оборудования по месячному плану",
        ["№", "Код оборудования", "Оборудование", "Доступно, ч", "Плановая загрузка, ч", "Плановое ТО, ч", "Резерв / перегруз, ч", "Загрузка, %", "Статус"],
        load_rows,
        numeric_columns={4, 5, 6, 7},
        percent_columns={8},
        wrap_columns={3, 9},
    )

    downtime_summary_rows = [
        ["Всего простоев, ч", summary.total_downtime_hours],
        ["Плановое ТО, ч", summary.planned_maintenance_hours],
        ["Внеплановые простои, ч", summary.unplanned_downtime_hours],
        ["Доля внеплановых, %", summary.unplanned_share_percent],
    ]
    current_row = write_summary_table(
        worksheet,
        current_row,
        "Сводка простоев за месяц",
        downtime_summary_rows,
        last_column=9,
    )

    category_rows = [
        [index, item.category, item.downtime_count, item.downtime_hours, item.share_percent]
        for index, item in enumerate(data.downtime_by_category or [], start=1)
    ]
    current_row = write_table(
        worksheet,
        current_row,
        "Внеплановые простои по категориям",
        ["№", "Категория", "Кол-во простоев", "Время, ч", "Доля, %"],
        category_rows,
        numeric_columns={3, 4},
        percent_columns={5},
        wrap_columns={2},
    )

    downtime_rows = [
        [
            index,
            item.equipment_code,
            item.equipment_name,
            item.reason_name,
            item.reason_category,
            item.downtime_count,
            item.downtime_hours,
        ]
        for index, item in enumerate(data.downtimes or [], start=1)
    ]
    current_row = write_table(
        worksheet,
        current_row,
        "Детализация внеплановых простоев",
        ["№", "Код оборудования", "Оборудование", "Причина", "Категория", "Кол-во", "Время, ч"],
        downtime_rows,
        numeric_columns={6, 7},
        wrap_columns={3, 4, 5},
    )

    worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    note_cell = worksheet.cell(
        row=current_row,
        column=1,
        value="Плановое ТО учтено в доступности оборудования. Внеплановые простои показаны отдельно как факт периода и не изменяют плановую доступность.",
    )
    note_cell.font = Font(size=10, italic=True, color="475569")
    note_cell.alignment = Alignment(wrap_text=True)


@router.get("/print", dependencies=[Depends(require_roles(*ANALYTICS_READ_ROLES))])
def print_production_analytics(month: str = Query(...)):
    month_start = parse_month_value(month)
    month_label = month_start.strftime("%Y-%m")

    try:
        output_data = get_monthly_output_analytics(month=month_label, only_with_deviations=False)
        equipment_data = get_equipment_monthly_analytics(month=month_label)
        workbook_bytes = create_production_analytics_print_workbook(
            month=month_label,
            output_data=output_data,
            equipment_data=equipment_data,
        )
        filename = f"Анализ_выпуска_{month_label}.xlsx"
        encoded_filename = quote(filename)
        return StreamingResponse(
            BytesIO(workbook_bytes),
            media_type=XLSX_MEDIA_TYPE,
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            },
        )
    except HTTPException:
        raise
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось сформировать печатную форму анализа выпуска.",
        ) from exc


@router.get(
    "/monthly-output",
    response_model=MonthlyOutputAnalyticsResponse,
    dependencies=[Depends(require_roles(*ANALYTICS_READ_ROLES))],
)
def get_monthly_output_analytics(
    month: str = Query(...),
    only_with_deviations: bool = Query(default=False),
):
    month_start = parse_month_value(month)
    month_end = get_next_month(month_start)
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                WITH selected_plan AS (
                    SELECT production_plan_id
                    FROM production_plans
                    WHERE plan_month = %s
                    LIMIT 1
                ),
                plan_items AS (
                    SELECT
                        ppl.nomenclature_id,
                        n.nomenclature_code AS item_code,
                        n.nomenclature_name AS item_name,
                        n.unit_of_measure AS unit_of_measure,
                        COALESCE(SUM(ppl.planned_qty), 0) AS planned_qty
                    FROM production_plan_lines AS ppl
                    INNER JOIN selected_plan AS sp ON sp.production_plan_id = ppl.production_plan_id
                    INNER JOIN nomenclature AS n ON n.nomenclature_id = ppl.nomenclature_id
                    GROUP BY
                        ppl.nomenclature_id,
                        n.nomenclature_code,
                        n.nomenclature_name,
                        n.unit_of_measure
                ),
                actual_items AS (
                    SELECT
                        pa.nomenclature_id,
                        n.nomenclature_code AS item_code,
                        n.nomenclature_name AS item_name,
                        n.unit_of_measure AS unit_of_measure,
                        COALESCE(SUM(pa.actual_qty), 0) AS actual_qty
                    FROM production_actuals AS pa
                    INNER JOIN nomenclature AS n ON n.nomenclature_id = pa.nomenclature_id
                    WHERE pa.actual_date >= %s
                      AND pa.actual_date < %s
                    GROUP BY
                        pa.nomenclature_id,
                        n.nomenclature_code,
                        n.nomenclature_name,
                        n.unit_of_measure
                )
                SELECT
                    COALESCE(plan_items.nomenclature_id, actual_items.nomenclature_id) AS nomenclature_id,
                    COALESCE(plan_items.item_code, actual_items.item_code) AS item_code,
                    COALESCE(plan_items.item_name, actual_items.item_name) AS item_name,
                    COALESCE(plan_items.unit_of_measure, actual_items.unit_of_measure) AS unit_of_measure,
                    COALESCE(plan_items.planned_qty, 0) AS planned_qty,
                    COALESCE(actual_items.actual_qty, 0) AS actual_qty
                FROM plan_items
                FULL OUTER JOIN actual_items
                    ON actual_items.nomenclature_id = plan_items.nomenclature_id
                ORDER BY
                    COALESCE(plan_items.item_code, actual_items.item_code) ASC,
                    COALESCE(plan_items.item_name, actual_items.item_name) ASC;
                """,
                (month_start, month_start, month_end),
            )
            rows = cursor.fetchall()

        items: list[MonthlyOutputAnalyticsItem] = []
        planned_qty_total = DECIMAL_ZERO
        actual_qty_total = DECIMAL_ZERO
        remaining_qty_total = DECIMAL_ZERO
        underproduced_items_count = 0
        overproduced_items_count = 0
        no_actual_items_count = 0
        no_plan_items_count = 0
        summary_by_unit_totals: dict[str, dict[str, Decimal]] = {}

        for row in rows:
            planned_qty = to_decimal(row["planned_qty"])
            actual_qty = to_decimal(row["actual_qty"])
            remaining_qty = planned_qty - actual_qty if actual_qty < planned_qty else DECIMAL_ZERO
            deviation_qty = actual_qty - planned_qty
            completion_percent: Decimal | None = None
            if planned_qty > DECIMAL_ZERO:
                completion_percent = (actual_qty / planned_qty) * Decimal("100")

            status, status_label = build_item_status(planned_qty, actual_qty)

            planned_qty_total += planned_qty
            actual_qty_total += actual_qty
            remaining_qty_total += remaining_qty
            unit = normalize_unit_of_measure(row.get("unit_of_measure"))
            unit_totals = summary_by_unit_totals.setdefault(
                unit,
                {
                    "planned_qty_total": DECIMAL_ZERO,
                    "actual_qty_total": DECIMAL_ZERO,
                    "remaining_qty_total": DECIMAL_ZERO,
                },
            )
            unit_totals["planned_qty_total"] += planned_qty
            unit_totals["actual_qty_total"] += actual_qty
            unit_totals["remaining_qty_total"] += remaining_qty

            if status == STATUS_IN_PROGRESS:
                underproduced_items_count += 1
            elif status == STATUS_OVERPRODUCED:
                overproduced_items_count += 1
            elif status == STATUS_NO_ACTUAL:
                underproduced_items_count += 1
                no_actual_items_count += 1
            elif status == STATUS_NO_PLAN:
                no_plan_items_count += 1

            items.append(
                MonthlyOutputAnalyticsItem(
                    nomenclature_id=int(row["nomenclature_id"]),
                    item_code=str(row["item_code"] or "").strip(),
                    item_name=str(row["item_name"] or "").strip(),
                    planned_qty=to_qty_float(planned_qty),
                    actual_qty=to_qty_float(actual_qty),
                    remaining_qty=to_qty_float(remaining_qty),
                    deviation_qty=to_qty_float(deviation_qty),
                    completion_percent=to_percent_float(completion_percent),
                    status=status,
                    status_label=status_label,
                )
            )

        filtered_items = items
        if only_with_deviations:
            filtered_items = [item for item in items if item.status != STATUS_COMPLETED]

        top_problem_items = sorted(
            (
                MonthlyOutputAnalyticsProblemItem(
                    nomenclature_id=item.nomenclature_id,
                    item_code=item.item_code,
                    item_name=item.item_name,
                    planned_qty=item.planned_qty,
                    actual_qty=item.actual_qty,
                    remaining_qty=item.remaining_qty,
                    completion_percent=item.completion_percent,
                    status=item.status,
                    status_label=item.status_label,
                )
                for item in items
                if item.status in {STATUS_NO_ACTUAL, STATUS_IN_PROGRESS}
            ),
            key=lambda item: (-item.remaining_qty, -item.planned_qty, item.item_code),
        )[:5]

        summary = build_empty_summary()
        if items:
            completion_percent_total = (
                (actual_qty_total / planned_qty_total) * Decimal("100")
                if planned_qty_total > DECIMAL_ZERO
                else DECIMAL_ZERO
            )
            summary = MonthlyOutputAnalyticsSummary(
                planned_qty_total=to_qty_float(planned_qty_total),
                actual_qty_total=to_qty_float(actual_qty_total),
                remaining_qty_total=to_qty_float(remaining_qty_total),
                completion_percent=to_percent_float(completion_percent_total) or 0.0,
                underproduced_items_count=underproduced_items_count,
                overproduced_items_count=overproduced_items_count,
                no_actual_items_count=no_actual_items_count,
                no_plan_items_count=no_plan_items_count,
            )

        summary_by_unit = [
            MonthlyOutputAnalyticsSummaryByUnit(
                unit=unit,
                planned_qty_total=to_qty_float(totals["planned_qty_total"]),
                actual_qty_total=to_qty_float(totals["actual_qty_total"]),
                remaining_qty_total=to_qty_float(totals["remaining_qty_total"]),
            )
            for unit, totals in sorted(summary_by_unit_totals.items(), key=lambda item: item[0])
        ]

        return MonthlyOutputAnalyticsResponse(
            month=month_start.strftime("%Y-%m"),
            date_from=month_start,
            date_to=month_end,
            summary=summary,
            summary_by_unit=summary_by_unit,
            top_problem_items=top_problem_items,
            items=filtered_items,
        )
    except HTTPException:
        raise
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось получить аналитику выпуска за месяц.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.get(
    "/capacity-monthly",
    response_model=EquipmentMonthlyAnalyticsResponse,
    dependencies=[Depends(require_roles(*ANALYTICS_READ_ROLES))],
)
@router.get(
    "/equipment-monthly",
    response_model=EquipmentMonthlyAnalyticsResponse,
    dependencies=[Depends(require_roles(*ANALYTICS_READ_ROLES))],
)
def get_equipment_monthly_analytics(month: str = Query(...)):
    month_start = parse_month_value(month)
    month_start, month_end, month_start_at, month_end_at = get_month_period_bounds(month_start)
    now_local = get_local_now()
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                SELECT
                    production_plan_week_id,
                    production_plan_id,
                    week_no,
                    week_start_date,
                    week_end_date,
                    status,
                    comment,
                    created_at,
                    updated_at
                FROM production_plan_weeks
                WHERE week_start_date < %s
                  AND week_end_date >= %s
                ORDER BY week_start_date ASC, production_plan_week_id ASC;
                """,
                (month_end, month_start),
            )
            week_rows = cursor.fetchall()

            equipment_totals: dict[int, dict[str, Any]] = {}

            for week_row in week_rows:
                week_id = int(week_row["production_plan_week_id"])
                week_lines = fetch_week_lines_for_equipment_analytics(cursor, week_id)
                equipment_availability, _ = build_equipment_availability(cursor, week_row, week_lines)

                capacity_gap_by_machine: dict[int, bool] = {}
                for line in week_lines:
                    machine_id_raw = line.get("machine_id")
                    route_step_equipment_id = line.get("route_step_equipment_id")
                    planned_qty = to_decimal(line.get("planned_qty"))
                    nominal_rate = to_decimal(line.get("nominal_rate"))
                    if machine_id_raw is None or route_step_equipment_id is None:
                        continue
                    if planned_qty <= DECIMAL_ZERO:
                        continue
                    if nominal_rate <= DECIMAL_ZERO:
                        capacity_gap_by_machine[int(machine_id_raw)] = True

                for item in equipment_availability:
                    equipment_id = int(item["machine_id"])
                    record = equipment_totals.setdefault(
                        equipment_id,
                        {
                            "equipment_id": equipment_id,
                            "equipment_code": str(item.get("machine_code") or "").strip(),
                            "equipment_name": str(item.get("machine_name") or "").strip(),
                            "available_minutes": 0,
                            "planned_load_minutes": DECIMAL_ZERO,
                            "maintenance_minutes": 0,
                            "has_capacity_gap": False,
                        },
                    )
                    record["equipment_code"] = str(item.get("machine_code") or record["equipment_code"] or "").strip()
                    record["equipment_name"] = str(item.get("machine_name") or record["equipment_name"] or "").strip()
                    record["available_minutes"] += int(item.get("available_minutes") or 0)
                    record["maintenance_minutes"] += int(item.get("maintenance_minutes") or 0)
                    record["planned_load_minutes"] += to_decimal(item.get("planned_load_minutes"))
                    record["has_capacity_gap"] = bool(record["has_capacity_gap"]) or capacity_gap_by_machine.get(equipment_id, False)

            equipment_load: list[EquipmentMonthlyLoadItem] = []
            load_percents: list[Decimal] = []
            overloaded_equipment_count = 0
            high_load_equipment_count = 0

            for record in sorted(
                equipment_totals.values(),
                key=lambda item: (item["equipment_code"], item["equipment_name"], item["equipment_id"]),
            ):
                available_minutes = int(record["available_minutes"])
                planned_load_minutes = to_decimal(record["planned_load_minutes"])
                has_capacity_gap = bool(record["has_capacity_gap"])
                status, status_label, warning, load_percent_decimal = build_equipment_status(
                    available_minutes=available_minutes,
                    planned_load_minutes=planned_load_minutes,
                    has_capacity_gap=has_capacity_gap,
                )

                if status == EQUIPMENT_STATUS_OVERLOADED:
                    overloaded_equipment_count += 1
                elif status == EQUIPMENT_STATUS_HIGH_LOAD:
                    high_load_equipment_count += 1

                if load_percent_decimal is not None:
                    load_percents.append(load_percent_decimal)

                planned_load_hours = None if has_capacity_gap else to_hours_float(planned_load_minutes)
                planned_maintenance_hours = to_hours_float(record.get("maintenance_minutes"))
                remaining_minutes = Decimal(available_minutes) - planned_load_minutes
                remaining_hours = to_hours_float(remaining_minutes)

                equipment_load.append(
                    EquipmentMonthlyLoadItem(
                        equipment_id=int(record["equipment_id"]),
                        equipment_code=str(record["equipment_code"] or "").strip(),
                        equipment_name=str(record["equipment_name"] or "").strip(),
                        available_hours=to_hours_float(available_minutes),
                        planned_load_hours=planned_load_hours,
                        planned_maintenance_hours=planned_maintenance_hours,
                        remaining_hours=remaining_hours,
                        load_percent=to_percent_float(load_percent_decimal),
                        status=status,
                        status_label=status_label,
                        warning=warning,
                    )
                )

            cursor.execute(
                """
                SELECT
                    ed.machine_id AS equipment_id,
                    m.machine_code AS equipment_code,
                    m.machine_name AS equipment_name,
                    ed.downtime_reason_id AS reason_id,
                    dr.reason_code,
                    dr.reason_name,
                    dr.reason_category,
                    ed.started_at,
                    ed.ended_at
                FROM equipment_downtimes AS ed
                INNER JOIN machines AS m ON m.machine_id = ed.machine_id
                INNER JOIN downtime_reasons AS dr ON dr.downtime_reason_id = ed.downtime_reason_id
                WHERE ed.started_at < %s
                  AND (ed.ended_at IS NULL OR ed.ended_at > %s)
                ORDER BY
                    m.machine_code ASC,
                    dr.reason_category ASC,
                    dr.reason_code ASC,
                    ed.started_at ASC;
                """,
                (month_end_at, month_start_at),
            )
            downtime_rows = cursor.fetchall()
            planned_maintenance_minutes_total = fetch_planned_maintenance_minutes_for_period(
                cursor=cursor,
                date_from=month_start_at,
                date_to=month_end_at,
            )

        downtime_aggregates: dict[tuple[int, int], dict[str, Any]] = {}
        downtime_category_aggregates: dict[str, dict[str, Any]] = {}
        unplanned_downtime_minutes_total = 0

        for row in downtime_rows:
            started_at = row.get("started_at")
            ended_at = row.get("ended_at")
            if not isinstance(started_at, datetime):
                continue

            effective_start = started_at if started_at >= month_start_at else month_start_at
            raw_end = ended_at if isinstance(ended_at, datetime) else now_local
            effective_end = raw_end if raw_end <= month_end_at else month_end_at
            duration_minutes = max(0, int((effective_end - effective_start).total_seconds() // 60))
            if duration_minutes <= 0:
                continue

            key = (int(row["equipment_id"]), int(row["reason_id"]))
            aggregate = downtime_aggregates.setdefault(
                key,
                {
                    "equipment_id": int(row["equipment_id"]),
                    "equipment_code": str(row.get("equipment_code") or "").strip(),
                    "equipment_name": str(row.get("equipment_name") or "").strip(),
                    "reason_id": int(row["reason_id"]),
                    "reason_code": str(row.get("reason_code") or "").strip() or None,
                    "reason_name": str(row.get("reason_name") or "").strip(),
                    "reason_category": str(row.get("reason_category") or "").strip(),
                    "downtime_count": 0,
                    "downtime_minutes": 0,
                },
            )
            aggregate["downtime_count"] += 1
            aggregate["downtime_minutes"] += duration_minutes

            category_name = str(row.get("reason_category") or "").strip() or "Без категории"
            category_aggregate = downtime_category_aggregates.setdefault(
                category_name,
                {
                    "category": category_name,
                    "downtime_count": 0,
                    "downtime_minutes": 0,
                },
            )
            category_aggregate["downtime_count"] += 1
            category_aggregate["downtime_minutes"] += duration_minutes
            unplanned_downtime_minutes_total += duration_minutes

        downtimes = sorted(
            [
                EquipmentMonthlyDowntimeItem(
                    equipment_id=item["equipment_id"],
                    equipment_code=item["equipment_code"],
                    equipment_name=item["equipment_name"],
                    reason_id=item["reason_id"],
                    reason_code=item["reason_code"],
                    reason_name=item["reason_name"],
                    reason_category=item["reason_category"],
                    downtime_count=int(item["downtime_count"]),
                    downtime_hours=to_hours_float(item["downtime_minutes"]),
                )
                for item in downtime_aggregates.values()
            ],
            key=lambda item: (-item.downtime_hours, item.equipment_code, item.reason_name, item.reason_id),
        )

        unplanned_downtime_hours_total = to_hours_float(unplanned_downtime_minutes_total)
        downtime_by_category = sorted(
            [
                EquipmentMonthlyDowntimeCategoryItem(
                    category=item["category"],
                    downtime_count=int(item["downtime_count"]),
                    downtime_hours=to_hours_float(item["downtime_minutes"]),
                    share_percent=round(
                        (to_hours_float(item["downtime_minutes"]) / unplanned_downtime_hours_total) * 100,
                        1,
                    )
                    if unplanned_downtime_hours_total > 0
                    else 0.0,
                )
                for item in downtime_category_aggregates.values()
            ],
            key=lambda item: (-item.downtime_hours, item.category),
        )

        average_load_percent = (
            (sum(load_percents, DECIMAL_ZERO) / Decimal(len(load_percents))).quantize(Decimal("0.1"))
            if load_percents
            else DECIMAL_ZERO
        )

        summary = build_empty_equipment_summary()
        maintenance_hours_total = to_hours_float(planned_maintenance_minutes_total)
        unplanned_hours_total = unplanned_downtime_hours_total
        total_downtime_hours = round(maintenance_hours_total + unplanned_hours_total, 1)
        unplanned_share_percent = (
            round((unplanned_hours_total / total_downtime_hours) * 100, 1)
            if total_downtime_hours > 0
            else 0.0
        )
        if equipment_load or downtimes or planned_maintenance_minutes_total > 0:
            summary = EquipmentMonthlyAnalyticsSummary(
                equipment_in_plan_count=len(equipment_load),
                average_load_percent=to_percent_float(average_load_percent) or 0.0,
                overloaded_equipment_count=overloaded_equipment_count,
                high_load_equipment_count=high_load_equipment_count,
                total_downtime_hours=total_downtime_hours,
                planned_maintenance_hours=maintenance_hours_total,
                unplanned_downtime_hours=unplanned_hours_total,
                unplanned_share_percent=unplanned_share_percent,
            )

        return EquipmentMonthlyAnalyticsResponse(
            month=month_start.strftime("%Y-%m"),
            date_from=month_start,
            date_to=month_end,
            summary=summary,
            equipment_load=equipment_load,
            downtime_by_category=downtime_by_category,
            downtimes=downtimes,
        )
    except HTTPException:
        raise
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось получить аналитику по оборудованию за месяц.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()
