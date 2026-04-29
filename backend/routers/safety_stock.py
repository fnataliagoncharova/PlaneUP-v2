from decimal import Decimal, InvalidOperation
from io import BytesIO
import re
from typing import Any

import psycopg2
from fastapi import APIRouter, File, Form, HTTPException, Path, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from psycopg2.errors import UniqueViolation
from psycopg2.extras import RealDictCursor

from db import get_connection
from schemas.safety_stock import (
    SafetyStockCreate,
    SafetyStockDeleteResponse,
    SafetyStockImportCommitResponse,
    SafetyStockImportCommitRow,
    SafetyStockImportMode,
    SafetyStockImportPreviewResponse,
    SafetyStockImportPreviewRow,
    SafetyStockRead,
    SafetyStockUpdate,
)


router = APIRouter(prefix="/safety-stock", tags=["safety_stock"])

IMPORT_MODE_UPSERT: SafetyStockImportMode = "upsert"

HEADER_FIELD_NOMENCLATURE_CODE = "nomenclature_code"
HEADER_FIELD_NOMENCLATURE_NAME = "nomenclature_name"
HEADER_FIELD_STOCK_QTY = "stock_qty"
HEADER_FIELD_UNIT_OF_MEASURE = "unit_of_measure"

REQUIRED_IMPORT_HEADERS = {
    HEADER_FIELD_NOMENCLATURE_CODE,
    HEADER_FIELD_NOMENCLATURE_NAME,
    HEADER_FIELD_STOCK_QTY,
    HEADER_FIELD_UNIT_OF_MEASURE,
}

HEADER_ALIASES: dict[str, set[str]] = {
    HEADER_FIELD_NOMENCLATURE_CODE: {"кодноменклатуры", "nomenclaturecode"},
    HEADER_FIELD_NOMENCLATURE_NAME: {"наименованиеноменклатуры", "nomenclaturename"},
    HEADER_FIELD_STOCK_QTY: {"страховойзапас", "количество", "stockqty"},
    HEADER_FIELD_UNIT_OF_MEASURE: {"единицаизмерения", "unitofmeasure"},
}


def normalize_header_name(value: object) -> str:
    if value is None:
        return ""

    normalized_value = str(value).strip().lower().replace("ё", "е")
    return re.sub(r"[\s_\-./\\]+", "", normalized_value)


def resolve_import_header(value: object) -> str | None:
    normalized_value = normalize_header_name(value)
    if not normalized_value:
        return None

    for field_name, aliases in HEADER_ALIASES.items():
        if normalized_value in aliases:
            return field_name

    return None


def normalize_import_mode(import_mode: str | None) -> SafetyStockImportMode:
    normalized_mode = (import_mode or IMPORT_MODE_UPSERT).strip().lower()
    if normalized_mode != IMPORT_MODE_UPSERT:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Поддерживается только режим upsert.",
        )

    return IMPORT_MODE_UPSERT


def normalize_code(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_name(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_unit_of_measure(value: object) -> tuple[str | None, str | None]:
    if value is None:
        return None, "Недопустимая единица измерения"

    raw_value = str(value).strip()
    if not raw_value:
        return None, "Недопустимая единица измерения"

    lowered_value = raw_value.lower()
    cyrillic_value = lowered_value.translate(str.maketrans({"m": "м", "p": "п"}))
    compact_value = re.sub(r"\s+", "", cyrillic_value)
    compact_without_dots = compact_value.replace(".", "")

    if compact_value in {"м²", "м2", "м^2", "m2"}:
        return "м²", None

    if compact_without_dots in {"мп", "мпог", "мпогон", "мпогонный"}:
        return "м.п.", None

    if compact_without_dots.startswith("м") and "пог" in compact_without_dots:
        return "м.п.", None

    if compact_without_dots in {"шт", "штука", "штуки", "pcs", "pc"}:
        return "шт", None

    if compact_without_dots in {"кг", "kg"}:
        return "кг", None

    if compact_without_dots in {"л", "литр", "литры", "l"}:
        return "л", None

    return None, "Недопустимая единица измерения"


def canonicalize_system_unit(value: object) -> str:
    normalized_value, _ = normalize_unit_of_measure(value)
    if normalized_value:
        return normalized_value
    return str(value).strip() if value is not None else ""


def normalize_stock_qty(value: object) -> tuple[Decimal | None, str | None]:
    if value is None:
        return None, "Пустой страховой запас"

    if isinstance(value, Decimal):
        decimal_value = value
    elif isinstance(value, (int, float)):
        decimal_value = Decimal(str(value))
    else:
        raw_text = str(value).strip().replace(" ", "")
        if not raw_text:
            return None, "Пустой страховой запас"

        raw_text = raw_text.replace(",", ".")
        try:
            decimal_value = Decimal(raw_text)
        except InvalidOperation:
            return None, "Некорректный страховой запас"

    if decimal_value < 0:
        return None, "Страховой запас не может быть отрицательным"

    return decimal_value.quantize(Decimal("0.001")), None


def validate_import_file(file: UploadFile, file_bytes: bytes) -> None:
    file_name = file.filename or ""
    if not file_name.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Поддерживается только формат .xlsx.",
        )

    if len(file_bytes) == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Файл пустой.",
        )


def read_import_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Не удалось прочитать Excel-файл. Проверьте формат .xlsx.",
        ) from exc

    try:
        sheet = workbook.active
        max_column = sheet.max_column or 0
        max_row = sheet.max_row or 0

        if max_column == 0 or max_row == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Файл пустой.",
            )

        header_indexes: dict[str, int] = {}
        for column_index in range(1, max_column + 1):
            header_name = resolve_import_header(sheet.cell(row=1, column=column_index).value)
            if header_name and header_name not in header_indexes:
                header_indexes[header_name] = column_index

        missing_headers = [
            header_name
            for header_name in REQUIRED_IMPORT_HEADERS
            if header_name not in header_indexes
        ]
        if missing_headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "Не найдены обязательные колонки шаблона: "
                    "Код номенклатуры, Наименование номенклатуры, Страховой запас, Единица измерения."
                ),
            )

        rows: list[dict[str, Any]] = []
        for row_index in range(2, max_row + 1):
            raw_code = sheet.cell(
                row=row_index,
                column=header_indexes[HEADER_FIELD_NOMENCLATURE_CODE],
            ).value
            raw_name = sheet.cell(
                row=row_index,
                column=header_indexes[HEADER_FIELD_NOMENCLATURE_NAME],
            ).value
            raw_stock_qty = sheet.cell(
                row=row_index,
                column=header_indexes[HEADER_FIELD_STOCK_QTY],
            ).value
            raw_unit_of_measure = sheet.cell(
                row=row_index,
                column=header_indexes[HEADER_FIELD_UNIT_OF_MEASURE],
            ).value

            is_empty_row = (
                normalize_code(raw_code) == ""
                and normalize_name(raw_name) == ""
                and (raw_stock_qty is None or str(raw_stock_qty).strip() == "")
                and (raw_unit_of_measure is None or str(raw_unit_of_measure).strip() == "")
            )
            if is_empty_row:
                continue

            normalized_code = normalize_code(raw_code)
            normalized_name = normalize_name(raw_name)
            normalized_stock_qty, stock_qty_error = normalize_stock_qty(raw_stock_qty)
            normalized_uom, unit_error = normalize_unit_of_measure(raw_unit_of_measure)

            row_errors: list[str] = []
            if not normalized_code:
                row_errors.append("Пустой код номенклатуры")

            if stock_qty_error:
                row_errors.append(stock_qty_error)

            if unit_error:
                row_errors.append(unit_error)

            rows.append(
                {
                    "row_no": row_index,
                    "nomenclature_code": normalized_code or None,
                    "nomenclature_code_key": normalized_code.upper() if normalized_code else None,
                    "nomenclature_name": normalized_name or None,
                    "stock_qty": normalized_stock_qty,
                    "unit_of_measure": normalized_uom,
                    "errors": row_errors,
                    "unit_normalized_from": (
                        str(raw_unit_of_measure).strip()
                        if raw_unit_of_measure is not None
                        and normalized_uom is not None
                        and str(raw_unit_of_measure).strip() != normalized_uom
                        else None
                    ),
                }
            )

        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Файл пустой.",
            )

        return rows
    finally:
        workbook.close()


def fetch_nomenclature_by_code(cursor: RealDictCursor) -> dict[str, dict[str, Any]]:
    cursor.execute(
        """
        SELECT
            nomenclature_id,
            nomenclature_code,
            nomenclature_name,
            unit_of_measure
        FROM nomenclature;
        """
    )
    rows = cursor.fetchall()
    return {str(row["nomenclature_code"]).strip().upper(): row for row in rows}


def fetch_existing_safety_stock_nomenclature_ids(
    cursor: RealDictCursor,
    nomenclature_ids: set[int],
) -> set[int]:
    if not nomenclature_ids:
        return set()

    cursor.execute(
        """
        SELECT nomenclature_id
        FROM safety_stock
        WHERE nomenclature_id = ANY(%s);
        """,
        (sorted(nomenclature_ids),),
    )
    rows = cursor.fetchall()
    return {row["nomenclature_id"] for row in rows}


def collect_lookup_context(
    rows: list[dict[str, Any]],
    nomenclature_by_code: dict[str, dict[str, Any]],
) -> set[int]:
    nomenclature_ids: set[int] = set()

    for row in rows:
        code_key = row["nomenclature_code_key"]
        if not code_key:
            continue

        nomenclature_row = nomenclature_by_code.get(code_key)
        if not nomenclature_row:
            continue

        row["nomenclature_id"] = nomenclature_row["nomenclature_id"]
        nomenclature_ids.add(nomenclature_row["nomenclature_id"])

    return nomenclature_ids


def build_preview(
    parsed_rows: list[dict[str, Any]],
    nomenclature_by_code: dict[str, dict[str, Any]],
    existing_safety_stock_nomenclature_ids: set[int],
) -> SafetyStockImportPreviewResponse:
    duplicate_counts: dict[str, int] = {}

    for row in parsed_rows:
        code_key = row["nomenclature_code_key"]
        if code_key:
            duplicate_counts[code_key] = duplicate_counts.get(code_key, 0) + 1

    preview_rows: list[SafetyStockImportPreviewRow] = []
    valid_rows = 0
    new_rows = 0
    update_rows = 0
    error_rows = 0

    for row in parsed_rows:
        row_errors = list(row["errors"])
        code_key = row["nomenclature_code_key"]
        nomenclature_row = nomenclature_by_code.get(code_key) if code_key else None

        if nomenclature_row is None and code_key:
            row_errors.append("Номенклатура не найдена")

        if nomenclature_row is not None and row["nomenclature_name"] is None:
            row["nomenclature_name"] = nomenclature_row["nomenclature_name"]

        if nomenclature_row is not None and row["unit_of_measure"] is not None:
            system_uom = canonicalize_system_unit(nomenclature_row["unit_of_measure"])
            if row["unit_of_measure"] != system_uom:
                row_errors.append("Единица измерения не совпадает с номенклатурой")

        if code_key is not None and duplicate_counts.get(code_key, 0) > 1:
            row_errors.append("Дубликат строки в файле")

        status_value = "error"
        can_import = False

        if not row_errors and nomenclature_row is not None:
            if nomenclature_row["nomenclature_id"] in existing_safety_stock_nomenclature_ids:
                status_value = "update"
                update_rows += 1
            else:
                status_value = "new"
                new_rows += 1
            can_import = True
            valid_rows += 1
        else:
            error_rows += 1

        preview_rows.append(
            SafetyStockImportPreviewRow(
                row_no=row["row_no"],
                nomenclature_code=row["nomenclature_code"],
                nomenclature_name=row["nomenclature_name"],
                stock_qty=row["stock_qty"],
                unit_of_measure=row["unit_of_measure"],
                status=status_value,  # type: ignore[arg-type]
                can_import=can_import,
                messages=row_errors,
                unit_normalized_from=row["unit_normalized_from"],
            )
        )

    return SafetyStockImportPreviewResponse(
        import_mode=IMPORT_MODE_UPSERT,
        total_rows=len(parsed_rows),
        valid_rows=valid_rows,
        new_rows=new_rows,
        update_rows=update_rows,
        error_rows=error_rows,
        rows=preview_rows,
    )


def create_template_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Страховой запас"

    sheet.append(
        [
            "Код номенклатуры",
            "Наименование номенклатуры",
            "Страховой запас",
            "Единица измерения",
        ]
    )
    sheet.append(
        [
            "NM-001",
            "Полотно ламинированное белое",
            "500",
            "м²",
        ]
    )

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output.read()


def upsert_safety_stock_row(
    cursor: RealDictCursor,
    *,
    nomenclature_id: int,
    stock_qty: Decimal,
) -> bool:
    cursor.execute(
        """
        INSERT INTO safety_stock (
            nomenclature_id,
            stock_qty
        )
        VALUES (%s, %s)
        ON CONFLICT (nomenclature_id)
        DO UPDATE SET
            stock_qty = EXCLUDED.stock_qty,
            updated_at = NOW()
        RETURNING (xmax = 0) AS inserted;
        """,
        (
            nomenclature_id,
            stock_qty,
        ),
    )
    row = cursor.fetchone()
    return bool(row and row["inserted"])


@router.get("", response_model=list[SafetyStockRead])
def list_safety_stock(
    nomenclature_id: int | None = Query(default=None),
):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            if nomenclature_id is None:
                cursor.execute(
                    """
                    SELECT
                        ss.safety_stock_id,
                        ss.nomenclature_id,
                        n.nomenclature_code,
                        n.nomenclature_name,
                        ss.stock_qty,
                        n.unit_of_measure
                    FROM safety_stock AS ss
                    INNER JOIN nomenclature AS n ON n.nomenclature_id = ss.nomenclature_id
                    ORDER BY n.nomenclature_code
                    LIMIT 300;
                    """
                )
            else:
                cursor.execute(
                    """
                    SELECT
                        ss.safety_stock_id,
                        ss.nomenclature_id,
                        n.nomenclature_code,
                        n.nomenclature_name,
                        ss.stock_qty,
                        n.unit_of_measure
                    FROM safety_stock AS ss
                    INNER JOIN nomenclature AS n ON n.nomenclature_id = ss.nomenclature_id
                    WHERE ss.nomenclature_id = %s
                    ORDER BY n.nomenclature_code;
                    """,
                    (nomenclature_id,),
                )
            rows = cursor.fetchall()

        return rows
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось получить страховой запас.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.post("", response_model=SafetyStockRead, status_code=status.HTTP_201_CREATED)
def create_safety_stock_item(payload: SafetyStockCreate):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                SELECT nomenclature_id
                FROM nomenclature
                WHERE nomenclature_id = %s;
                """,
                (payload.nomenclature_id,),
            )
            nomenclature_row = cursor.fetchone()
            if nomenclature_row is None:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Номенклатура не найдена.",
                )

            cursor.execute(
                """
                INSERT INTO safety_stock (
                    nomenclature_id,
                    stock_qty
                )
                VALUES (%s, %s)
                RETURNING safety_stock_id;
                """,
                (
                    payload.nomenclature_id,
                    payload.stock_qty,
                ),
            )
            created_row = cursor.fetchone()

            cursor.execute(
                """
                SELECT
                    ss.safety_stock_id,
                    ss.nomenclature_id,
                    n.nomenclature_code,
                    n.nomenclature_name,
                    ss.stock_qty,
                    n.unit_of_measure
                FROM safety_stock AS ss
                INNER JOIN nomenclature AS n ON n.nomenclature_id = ss.nomenclature_id
                WHERE ss.safety_stock_id = %s;
                """,
                (created_row["safety_stock_id"],),
            )
            response_row = cursor.fetchone()

        connection.commit()
        return response_row
    except HTTPException:
        if connection is not None:
            connection.rollback()
        raise
    except UniqueViolation as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Позиция уже есть в страховом запасе.",
        ) from exc
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось создать строку страхового запаса.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.put("/{safety_stock_id}", response_model=SafetyStockRead)
def update_safety_stock_item(
    payload: SafetyStockUpdate,
    safety_stock_id: int = Path(..., gt=0),
):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                UPDATE safety_stock
                SET
                    stock_qty = %s,
                    updated_at = NOW()
                WHERE safety_stock_id = %s
                RETURNING safety_stock_id;
                """,
                (
                    payload.stock_qty,
                    safety_stock_id,
                ),
            )
            updated_row = cursor.fetchone()

            if updated_row is None:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Строка страхового запаса не найдена.",
                )

            cursor.execute(
                """
                SELECT
                    ss.safety_stock_id,
                    ss.nomenclature_id,
                    n.nomenclature_code,
                    n.nomenclature_name,
                    ss.stock_qty,
                    n.unit_of_measure
                FROM safety_stock AS ss
                INNER JOIN nomenclature AS n ON n.nomenclature_id = ss.nomenclature_id
                WHERE ss.safety_stock_id = %s;
                """,
                (safety_stock_id,),
            )
            response_row = cursor.fetchone()

        connection.commit()
        return response_row
    except HTTPException:
        if connection is not None:
            connection.rollback()
        raise
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось обновить строку страхового запаса.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.delete("/{safety_stock_id}", response_model=SafetyStockDeleteResponse)
def delete_safety_stock_item(safety_stock_id: int = Path(..., gt=0)):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                DELETE FROM safety_stock
                WHERE safety_stock_id = %s
                RETURNING safety_stock_id;
                """,
                (safety_stock_id,),
            )
            deleted_row = cursor.fetchone()

            if deleted_row is None:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Строка страхового запаса не найдена.",
                )

        connection.commit()
        return SafetyStockDeleteResponse(
            safety_stock_id=safety_stock_id,
            message="Строка страхового запаса удалена.",
        )
    except HTTPException:
        if connection is not None:
            connection.rollback()
        raise
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось удалить строку страхового запаса.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.get("/import/template")
def download_safety_stock_import_template():
    template_content = create_template_workbook()
    return StreamingResponse(
        BytesIO(template_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="safety_stock_import_template.xlsx"',
        },
    )


@router.post("/import/preview", response_model=SafetyStockImportPreviewResponse)
async def preview_safety_stock_import(
    file: UploadFile = File(...),
    import_mode: str = Form(IMPORT_MODE_UPSERT),
):
    normalize_import_mode(import_mode)
    file_bytes = await file.read()
    validate_import_file(file, file_bytes)
    parsed_rows = read_import_rows(file_bytes)

    connection = None
    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            nomenclature_by_code = fetch_nomenclature_by_code(cursor)
            nomenclature_ids = collect_lookup_context(parsed_rows, nomenclature_by_code)
            existing_safety_stock_nomenclature_ids = fetch_existing_safety_stock_nomenclature_ids(
                cursor,
                nomenclature_ids=nomenclature_ids,
            )

        return build_preview(
            parsed_rows=parsed_rows,
            nomenclature_by_code=nomenclature_by_code,
            existing_safety_stock_nomenclature_ids=existing_safety_stock_nomenclature_ids,
        )
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось подготовить предпросмотр импорта страхового запаса.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.post("/import/commit", response_model=SafetyStockImportCommitResponse)
async def commit_safety_stock_import(
    file: UploadFile = File(...),
    import_mode: str = Form(IMPORT_MODE_UPSERT),
):
    normalize_import_mode(import_mode)
    file_bytes = await file.read()
    validate_import_file(file, file_bytes)
    parsed_rows = read_import_rows(file_bytes)

    connection = None
    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            nomenclature_by_code = fetch_nomenclature_by_code(cursor)
            nomenclature_ids = collect_lookup_context(parsed_rows, nomenclature_by_code)
            existing_safety_stock_nomenclature_ids = fetch_existing_safety_stock_nomenclature_ids(
                cursor,
                nomenclature_ids=nomenclature_ids,
            )
            preview_response = build_preview(
                parsed_rows=parsed_rows,
                nomenclature_by_code=nomenclature_by_code,
                existing_safety_stock_nomenclature_ids=existing_safety_stock_nomenclature_ids,
            )

            commit_rows: list[SafetyStockImportCommitRow] = []
            created_count = 0
            updated_count = 0
            skipped_count = 0
            error_count = 0

            for preview_row in preview_response.rows:
                if not preview_row.can_import:
                    skipped_count += 1
                    error_count += 1
                    commit_rows.append(
                        SafetyStockImportCommitRow(
                            row_no=preview_row.row_no,
                            nomenclature_code=preview_row.nomenclature_code,
                            status="error",
                            message=(
                                preview_row.messages[0]
                                if preview_row.messages
                                else "Строка не прошла валидацию"
                            ),
                        )
                    )
                    continue

                code_key = (preview_row.nomenclature_code or "").strip().upper()
                nomenclature_row = nomenclature_by_code.get(code_key)
                if nomenclature_row is None or preview_row.stock_qty is None:
                    skipped_count += 1
                    error_count += 1
                    commit_rows.append(
                        SafetyStockImportCommitRow(
                            row_no=preview_row.row_no,
                            nomenclature_code=preview_row.nomenclature_code,
                            status="error",
                            message="Строка не прошла повторную проверку",
                        )
                    )
                    continue

                inserted = upsert_safety_stock_row(
                    cursor,
                    nomenclature_id=nomenclature_row["nomenclature_id"],
                    stock_qty=preview_row.stock_qty,
                )
                if inserted:
                    created_count += 1
                    commit_rows.append(
                        SafetyStockImportCommitRow(
                            row_no=preview_row.row_no,
                            nomenclature_code=preview_row.nomenclature_code,
                            status="created",
                            message="Создано",
                        )
                    )
                else:
                    updated_count += 1
                    commit_rows.append(
                        SafetyStockImportCommitRow(
                            row_no=preview_row.row_no,
                            nomenclature_code=preview_row.nomenclature_code,
                            status="updated",
                            message="Обновлено",
                        )
                    )

        connection.commit()
        return SafetyStockImportCommitResponse(
            import_mode=IMPORT_MODE_UPSERT,
            total_rows=preview_response.total_rows,
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=skipped_count,
            error_count=error_count,
            rows=commit_rows,
        )
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось выполнить импорт страхового запаса.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()
