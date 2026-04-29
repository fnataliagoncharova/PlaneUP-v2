from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re
from typing import Any

import psycopg2
from fastapi import APIRouter, File, Form, HTTPException, Path, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel
from psycopg2.errors import UniqueViolation
from psycopg2.extras import RealDictCursor

from db import get_connection
from schemas.inventory_balance import (
    InventoryBalanceCreate,
    InventoryBalanceDeleteResponse,
    InventoryBalanceImportCommitResponse,
    InventoryBalanceImportCommitRow,
    InventoryBalanceImportMode,
    InventoryBalanceImportPreviewResponse,
    InventoryBalanceImportPreviewRow,
    InventoryBalanceRead,
    InventoryBalanceUpdate,
)


router = APIRouter(prefix="/inventory-balance", tags=["inventory_balance"])

IMPORT_MODE_UPSERT: InventoryBalanceImportMode = "upsert"

HEADER_FIELD_AS_OF_DATE = "as_of_date"
HEADER_FIELD_NOMENCLATURE_CODE = "nomenclature_code"
HEADER_FIELD_NOMENCLATURE_NAME = "nomenclature_name"
HEADER_FIELD_AVAILABLE_QTY = "available_qty"
HEADER_FIELD_UNIT_OF_MEASURE = "unit_of_measure"

REQUIRED_IMPORT_HEADERS = {
    HEADER_FIELD_AS_OF_DATE,
    HEADER_FIELD_NOMENCLATURE_CODE,
    HEADER_FIELD_NOMENCLATURE_NAME,
    HEADER_FIELD_AVAILABLE_QTY,
    HEADER_FIELD_UNIT_OF_MEASURE,
}

HEADER_ALIASES: dict[str, set[str]] = {
    HEADER_FIELD_AS_OF_DATE: {"датаостатка", "asofdate"},
    HEADER_FIELD_NOMENCLATURE_CODE: {"кодноменклатуры", "nomenclaturecode"},
    HEADER_FIELD_NOMENCLATURE_NAME: {"наименованиеноменклатуры", "nomenclaturename"},
    HEADER_FIELD_AVAILABLE_QTY: {"доступныйостаток", "количество", "availableqty"},
    HEADER_FIELD_UNIT_OF_MEASURE: {"единицаизмерения", "unitofmeasure"},
}


def normalize_header_name(value: object) -> str:
    if value is None:
        return ""

    normalized_value = str(value).strip().lower().replace("ё", "е")
    return re.sub(r"[\s_\-./\\]+", "", normalized_value)


def resolve_import_header(value: object) -> str | None:
    normalized_value = normalize_header_name(value)
    if not normalized_value:
        return None

    for field_name, aliases in HEADER_ALIASES.items():
        if normalized_value in aliases:
            return field_name

    return None


def normalize_import_mode(import_mode: str | None) -> InventoryBalanceImportMode:
    normalized_mode = (import_mode or IMPORT_MODE_UPSERT).strip().lower()
    if normalized_mode != IMPORT_MODE_UPSERT:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Поддерживается только режим upsert.",
        )

    return IMPORT_MODE_UPSERT


def normalize_code(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_name(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_unit_of_measure(value: object) -> tuple[str | None, str | None]:
    if value is None:
        return None, "Недопустимая единица измерения"

    raw_value = str(value).strip()
    if not raw_value:
        return None, "Недопустимая единица измерения"

    lowered_value = raw_value.lower()
    cyrillic_value = lowered_value.translate(str.maketrans({"m": "м", "p": "п"}))
    compact_value = re.sub(r"\s+", "", cyrillic_value)
    compact_without_dots = compact_value.replace(".", "")

    if compact_value in {"м²", "м2", "м^2", "m2"}:
        return "м²", None

    if compact_without_dots in {"мп", "мпог", "мпогон", "мпогонный"}:
        return "м.п.", None

    if compact_without_dots.startswith("м") and "пог" in compact_without_dots:
        return "м.п.", None

    if compact_without_dots in {"шт", "штука", "штуки", "pcs", "pc"}:
        return "шт", None

    if compact_without_dots in {"кг", "kg"}:
        return "кг", None

    if compact_without_dots in {"л", "литр", "литры", "l"}:
        return "л", None

    return None, "Недопустимая единица измерения"


def canonicalize_system_unit(value: object) -> str:
    normalized_value, _ = normalize_unit_of_measure(value)
    if normalized_value:
        return normalized_value
    return str(value).strip() if value is not None else ""


def normalize_available_qty(value: object) -> tuple[Decimal | None, str | None]:
    if value is None:
        return None, "Пустой доступный остаток"

    if isinstance(value, Decimal):
        decimal_value = value
    elif isinstance(value, (int, float)):
        decimal_value = Decimal(str(value))
    else:
        raw_text = str(value).strip().replace(" ", "")
        if not raw_text:
            return None, "Пустой доступный остаток"

        raw_text = raw_text.replace(",", ".")
        try:
            decimal_value = Decimal(raw_text)
        except InvalidOperation:
            return None, "Некорректный доступный остаток"

    if decimal_value < 0:
        return None, "Доступный остаток не может быть отрицательным"

    return decimal_value.quantize(Decimal("0.001")), None


def normalize_as_of_date(value: object) -> tuple[date | None, str | None]:
    if value is None:
        return None, "Пустая дата остатка"

    if isinstance(value, datetime):
        return value.date(), None

    if isinstance(value, date):
        return value, None

    if isinstance(value, (int, float)):
        try:
            excel_datetime = from_excel(value)
            if isinstance(excel_datetime, datetime):
                return excel_datetime.date(), None
            if isinstance(excel_datetime, date):
                return excel_datetime, None
        except Exception:
            return None, "Некорректная дата остатка"

        return None, "Некорректная дата остатка"

    raw_value = str(value).strip()
    if not raw_value:
        return None, "Пустая дата остатка"

    date_formats = [
        "%Y-%m-%d",
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d-%m-%Y",
    ]
    for date_format in date_formats:
        try:
            return datetime.strptime(raw_value, date_format).date(), None
        except ValueError:
            continue

    return None, "Некорректная дата остатка"


def validate_import_file(file: UploadFile, file_bytes: bytes) -> None:
    file_name = file.filename or ""
    if not file_name.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Поддерживается только формат .xlsx.",
        )

    if len(file_bytes) == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Файл пустой.",
        )


def read_import_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Не удалось прочитать Excel-файл. Проверьте формат .xlsx.",
        ) from exc

    try:
        sheet = workbook.active
        max_column = sheet.max_column or 0
        max_row = sheet.max_row or 0

        if max_column == 0 or max_row == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Файл пустой.",
            )

        header_indexes: dict[str, int] = {}
        for column_index in range(1, max_column + 1):
            header_name = resolve_import_header(sheet.cell(row=1, column=column_index).value)
            if header_name and header_name not in header_indexes:
                header_indexes[header_name] = column_index

        missing_headers = [
            header_name
            for header_name in REQUIRED_IMPORT_HEADERS
            if header_name not in header_indexes
        ]
        if missing_headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "Не найдены обязательные колонки шаблона: "
                    "Дата остатка, Код номенклатуры, Наименование номенклатуры, Доступный остаток, Единица измерения."
                ),
            )

        rows: list[dict[str, Any]] = []
        for row_index in range(2, max_row + 1):
            raw_as_of_date = sheet.cell(
                row=row_index,
                column=header_indexes[HEADER_FIELD_AS_OF_DATE],
            ).value
            raw_code = sheet.cell(
                row=row_index,
                column=header_indexes[HEADER_FIELD_NOMENCLATURE_CODE],
            ).value
            raw_name = sheet.cell(
                row=row_index,
                column=header_indexes[HEADER_FIELD_NOMENCLATURE_NAME],
            ).value
            raw_available_qty = sheet.cell(
                row=row_index,
                column=header_indexes[HEADER_FIELD_AVAILABLE_QTY],
            ).value
            raw_unit_of_measure = sheet.cell(
                row=row_index,
                column=header_indexes[HEADER_FIELD_UNIT_OF_MEASURE],
            ).value

            is_empty_row = (
                (raw_as_of_date is None or str(raw_as_of_date).strip() == "")
                and normalize_code(raw_code) == ""
                and normalize_name(raw_name) == ""
                and (raw_available_qty is None or str(raw_available_qty).strip() == "")
                and (raw_unit_of_measure is None or str(raw_unit_of_measure).strip() == "")
            )
            if is_empty_row:
                continue

            normalized_as_of_date, as_of_date_error = normalize_as_of_date(raw_as_of_date)
            normalized_code = normalize_code(raw_code)
            normalized_name = normalize_name(raw_name)
            normalized_available_qty, available_qty_error = normalize_available_qty(raw_available_qty)
            normalized_uom, unit_error = normalize_unit_of_measure(raw_unit_of_measure)

            row_errors: list[str] = []
            if as_of_date_error:
                row_errors.append(as_of_date_error)

            if not normalized_code:
                row_errors.append("Пустой код номенклатуры")

            if available_qty_error:
                row_errors.append(available_qty_error)

            if unit_error:
                row_errors.append(unit_error)

            rows.append(
                {
                    "row_no": row_index,
                    "as_of_date": normalized_as_of_date,
                    "nomenclature_code": normalized_code or None,
                    "nomenclature_code_key": normalized_code.upper() if normalized_code else None,
                    "nomenclature_name": normalized_name or None,
                    "available_qty": normalized_available_qty,
                    "unit_of_measure": normalized_uom,
                    "errors": row_errors,
                    "unit_normalized_from": (
                        str(raw_unit_of_measure).strip()
                        if raw_unit_of_measure is not None
                        and normalized_uom is not None
                        and str(raw_unit_of_measure).strip() != normalized_uom
                        else None
                    ),
                }
            )

        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Файл пустой.",
            )

        return rows
    finally:
        workbook.close()


def fetch_nomenclature_by_code(cursor: RealDictCursor) -> dict[str, dict[str, Any]]:
    cursor.execute(
        """
        SELECT
            nomenclature_id,
            nomenclature_code,
            nomenclature_name,
            unit_of_measure
        FROM nomenclature;
        """
    )
    rows = cursor.fetchall()
    return {str(row["nomenclature_code"]).strip().upper(): row for row in rows}


def fetch_existing_inventory_balance_keys(
    cursor: RealDictCursor,
    as_of_dates: set[date],
    nomenclature_ids: set[int],
) -> set[tuple[date, int]]:
    if not as_of_dates or not nomenclature_ids:
        return set()

    cursor.execute(
        """
        SELECT as_of_date, nomenclature_id
        FROM inventory_balance
        WHERE as_of_date = ANY(%s)
          AND nomenclature_id = ANY(%s);
        """,
        (sorted(as_of_dates), sorted(nomenclature_ids)),
    )
    rows = cursor.fetchall()
    return {(row["as_of_date"], row["nomenclature_id"]) for row in rows}


def collect_lookup_context(
    rows: list[dict[str, Any]],
    nomenclature_by_code: dict[str, dict[str, Any]],
) -> tuple[set[date], set[int]]:
    as_of_dates: set[date] = set()
    nomenclature_ids: set[int] = set()

    for row in rows:
        code_key = row["nomenclature_code_key"]
        if not code_key:
            continue

        nomenclature_row = nomenclature_by_code.get(code_key)
        if not nomenclature_row:
            continue

        row["nomenclature_id"] = nomenclature_row["nomenclature_id"]
        if row["as_of_date"] is not None:
            as_of_dates.add(row["as_of_date"])
            nomenclature_ids.add(nomenclature_row["nomenclature_id"])

    return as_of_dates, nomenclature_ids


def build_preview(
    parsed_rows: list[dict[str, Any]],
    nomenclature_by_code: dict[str, dict[str, Any]],
    existing_inventory_balance_keys: set[tuple[date, int]],
) -> InventoryBalanceImportPreviewResponse:
    duplicate_counts: dict[tuple[date, str], int] = {}

    for row in parsed_rows:
        code_key = row["nomenclature_code_key"]
        row_as_of_date = row["as_of_date"]
        if row_as_of_date is not None and code_key:
            duplicate_key = (row_as_of_date, code_key)
            duplicate_counts[duplicate_key] = duplicate_counts.get(duplicate_key, 0) + 1

    preview_rows: list[InventoryBalanceImportPreviewRow] = []
    valid_rows = 0
    new_rows = 0
    update_rows = 0
    error_rows = 0

    for row in parsed_rows:
        row_errors = list(row["errors"])
        code_key = row["nomenclature_code_key"]
        nomenclature_row = nomenclature_by_code.get(code_key) if code_key else None

        if nomenclature_row is None and code_key:
            row_errors.append("Номенклатура не найдена")

        if nomenclature_row is not None and row["nomenclature_name"] is None:
            row["nomenclature_name"] = nomenclature_row["nomenclature_name"]

        if nomenclature_row is not None and row["unit_of_measure"] is not None:
            system_uom = canonicalize_system_unit(nomenclature_row["unit_of_measure"])
            if row["unit_of_measure"] != system_uom:
                row_errors.append("Единица измерения не совпадает с номенклатурой")

        duplicate_key = (
            row["as_of_date"],
            code_key,
        )
        if (
            row["as_of_date"] is not None
            and code_key is not None
            and duplicate_counts.get(duplicate_key, 0) > 1
        ):
            row_errors.append("Дубликат строки в файле")

        status_value = "error"
        can_import = False

        if not row_errors and nomenclature_row is not None:
            record_key = (
                row["as_of_date"],
                nomenclature_row["nomenclature_id"],
            )
            if record_key in existing_inventory_balance_keys:
                status_value = "update"
                update_rows += 1
            else:
                status_value = "new"
                new_rows += 1
            can_import = True
            valid_rows += 1
        else:
            error_rows += 1

        preview_rows.append(
            InventoryBalanceImportPreviewRow(
                row_no=row["row_no"],
                as_of_date=row["as_of_date"],
                nomenclature_code=row["nomenclature_code"],
                nomenclature_name=row["nomenclature_name"],
                available_qty=row["available_qty"],
                unit_of_measure=row["unit_of_measure"],
                status=status_value,  # type: ignore[arg-type]
                can_import=can_import,
                messages=row_errors,
                unit_normalized_from=row["unit_normalized_from"],
            )
        )

    return InventoryBalanceImportPreviewResponse(
        import_mode=IMPORT_MODE_UPSERT,
        total_rows=len(parsed_rows),
        valid_rows=valid_rows,
        new_rows=new_rows,
        update_rows=update_rows,
        error_rows=error_rows,
        rows=preview_rows,
    )


def create_template_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Остатки"

    sheet.append(
        [
            "Дата остатка",
            "Код номенклатуры",
            "Наименование номенклатуры",
            "Доступный остаток",
            "Единица измерения",
        ]
    )
    sheet.append(
        [
            "2026-05-01",
            "NM-001",
            "Полотно ламинированное белое",
            "2500",
            "м²",
        ]
    )

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output.read()


def upsert_inventory_balance_row(
    cursor: RealDictCursor,
    *,
    as_of_date: date,
    nomenclature_id: int,
    available_qty: Decimal,
) -> bool:
    cursor.execute(
        """
        INSERT INTO inventory_balance (
            as_of_date,
            nomenclature_id,
            available_qty
        )
        VALUES (%s, %s, %s)
        ON CONFLICT (as_of_date, nomenclature_id)
        DO UPDATE SET
            available_qty = EXCLUDED.available_qty,
            updated_at = NOW()
        RETURNING (xmax = 0) AS inserted;
        """,
        (
            as_of_date,
            nomenclature_id,
            available_qty,
        ),
    )
    row = cursor.fetchone()
    return bool(row and row["inserted"])


@router.get("", response_model=list[InventoryBalanceRead])
def list_inventory_balance(
    as_of_date: date | None = Query(default=None),
):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            if as_of_date is None:
                cursor.execute(
                    """
                    SELECT
                        ib.balance_id,
                        ib.as_of_date,
                        ib.nomenclature_id,
                        n.nomenclature_code,
                        n.nomenclature_name,
                        ib.available_qty,
                        n.unit_of_measure
                    FROM inventory_balance AS ib
                    INNER JOIN nomenclature AS n ON n.nomenclature_id = ib.nomenclature_id
                    ORDER BY ib.as_of_date DESC, n.nomenclature_code
                    LIMIT 300;
                    """
                )
            else:
                cursor.execute(
                    """
                    SELECT
                        ib.balance_id,
                        ib.as_of_date,
                        ib.nomenclature_id,
                        n.nomenclature_code,
                        n.nomenclature_name,
                        ib.available_qty,
                        n.unit_of_measure
                    FROM inventory_balance AS ib
                    INNER JOIN nomenclature AS n ON n.nomenclature_id = ib.nomenclature_id
                    WHERE ib.as_of_date = %s
                    ORDER BY n.nomenclature_code;
                    """,
                    (as_of_date,),
                )
            rows = cursor.fetchall()

        return rows
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось получить остатки.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.get("/dates", response_model=list[date])
def list_inventory_balance_dates():
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                SELECT DISTINCT as_of_date
                FROM inventory_balance
                ORDER BY as_of_date DESC;
                """
            )
            rows = cursor.fetchall()

        return [row["as_of_date"] for row in rows if row.get("as_of_date") is not None]
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось получить даты остатков.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.post("", response_model=InventoryBalanceRead, status_code=status.HTTP_201_CREATED)
def create_inventory_balance_item(payload: InventoryBalanceCreate):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                SELECT nomenclature_id
                FROM nomenclature
                WHERE nomenclature_id = %s;
                """,
                (payload.nomenclature_id,),
            )
            nomenclature_row = cursor.fetchone()
            if nomenclature_row is None:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Номенклатура не найдена.",
                )

            cursor.execute(
                """
                INSERT INTO inventory_balance (
                    as_of_date,
                    nomenclature_id,
                    available_qty
                )
                VALUES (%s, %s, %s)
                RETURNING balance_id;
                """,
                (
                    payload.as_of_date,
                    payload.nomenclature_id,
                    payload.available_qty,
                ),
            )
            created_row = cursor.fetchone()

            cursor.execute(
                """
                SELECT
                    ib.balance_id,
                    ib.as_of_date,
                    ib.nomenclature_id,
                    n.nomenclature_code,
                    n.nomenclature_name,
                    ib.available_qty,
                    n.unit_of_measure
                FROM inventory_balance AS ib
                INNER JOIN nomenclature AS n ON n.nomenclature_id = ib.nomenclature_id
                WHERE ib.balance_id = %s;
                """,
                (created_row["balance_id"],),
            )
            response_row = cursor.fetchone()

        connection.commit()
        return response_row
    except HTTPException:
        if connection is not None:
            connection.rollback()
        raise
    except UniqueViolation as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Позиция уже есть в остатках на выбранную дату.",
        ) from exc
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось создать строку остатков.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.put("/{balance_id}", response_model=InventoryBalanceRead)
def update_inventory_balance_item(
    payload: InventoryBalanceUpdate,
    balance_id: int = Path(..., gt=0),
):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                UPDATE inventory_balance
                SET
                    available_qty = %s,
                    updated_at = NOW()
                WHERE balance_id = %s
                RETURNING balance_id;
                """,
                (
                    payload.available_qty,
                    balance_id,
                ),
            )
            updated_row = cursor.fetchone()

            if updated_row is None:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Строка остатков не найдена.",
                )

            cursor.execute(
                """
                SELECT
                    ib.balance_id,
                    ib.as_of_date,
                    ib.nomenclature_id,
                    n.nomenclature_code,
                    n.nomenclature_name,
                    ib.available_qty,
                    n.unit_of_measure
                FROM inventory_balance AS ib
                INNER JOIN nomenclature AS n ON n.nomenclature_id = ib.nomenclature_id
                WHERE ib.balance_id = %s;
                """,
                (balance_id,),
            )
            response_row = cursor.fetchone()

        connection.commit()
        return response_row
    except HTTPException:
        if connection is not None:
            connection.rollback()
        raise
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось обновить строку остатков.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.delete("/{balance_id}", response_model=InventoryBalanceDeleteResponse)
def delete_inventory_balance_item(balance_id: int = Path(..., gt=0)):
    connection = None

    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                DELETE FROM inventory_balance
                WHERE balance_id = %s
                RETURNING balance_id;
                """,
                (balance_id,),
            )
            deleted_row = cursor.fetchone()

            if deleted_row is None:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Строка остатков не найдена.",
                )

        connection.commit()
        return InventoryBalanceDeleteResponse(
            balance_id=balance_id,
            message="Строка остатков удалена.",
        )
    except HTTPException:
        if connection is not None:
            connection.rollback()
        raise
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось удалить строку остатков.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.get("/import/template")
def download_inventory_balance_import_template():
    template_content = create_template_workbook()
    return StreamingResponse(
        BytesIO(template_content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="inventory_balance_import_template.xlsx"',
        },
    )


@router.post("/import/preview", response_model=InventoryBalanceImportPreviewResponse)
async def preview_inventory_balance_import(
    file: UploadFile = File(...),
    import_mode: str = Form(IMPORT_MODE_UPSERT),
):
    normalize_import_mode(import_mode)
    file_bytes = await file.read()
    validate_import_file(file, file_bytes)
    parsed_rows = read_import_rows(file_bytes)

    connection = None
    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            nomenclature_by_code = fetch_nomenclature_by_code(cursor)
            as_of_dates, nomenclature_ids = collect_lookup_context(parsed_rows, nomenclature_by_code)
            existing_inventory_balance_keys = fetch_existing_inventory_balance_keys(
                cursor,
                as_of_dates=as_of_dates,
                nomenclature_ids=nomenclature_ids,
            )

        return build_preview(
            parsed_rows=parsed_rows,
            nomenclature_by_code=nomenclature_by_code,
            existing_inventory_balance_keys=existing_inventory_balance_keys,
        )
    except psycopg2.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось подготовить предпросмотр импорта остатков.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()


@router.post("/import/commit", response_model=InventoryBalanceImportCommitResponse)
async def commit_inventory_balance_import(
    file: UploadFile = File(...),
    import_mode: str = Form(IMPORT_MODE_UPSERT),
):
    normalize_import_mode(import_mode)
    file_bytes = await file.read()
    validate_import_file(file, file_bytes)
    parsed_rows = read_import_rows(file_bytes)

    connection = None
    try:
        connection = get_connection()
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            nomenclature_by_code = fetch_nomenclature_by_code(cursor)
            as_of_dates, nomenclature_ids = collect_lookup_context(parsed_rows, nomenclature_by_code)
            existing_inventory_balance_keys = fetch_existing_inventory_balance_keys(
                cursor,
                as_of_dates=as_of_dates,
                nomenclature_ids=nomenclature_ids,
            )
            preview_response = build_preview(
                parsed_rows=parsed_rows,
                nomenclature_by_code=nomenclature_by_code,
                existing_inventory_balance_keys=existing_inventory_balance_keys,
            )

            commit_rows: list[InventoryBalanceImportCommitRow] = []
            created_count = 0
            updated_count = 0
            skipped_count = 0
            error_count = 0

            for preview_row in preview_response.rows:
                if not preview_row.can_import:
                    skipped_count += 1
                    error_count += 1
                    commit_rows.append(
                        InventoryBalanceImportCommitRow(
                            row_no=preview_row.row_no,
                            as_of_date=preview_row.as_of_date,
                            nomenclature_code=preview_row.nomenclature_code,
                            status="error",
                            message=(
                                preview_row.messages[0]
                                if preview_row.messages
                                else "Строка не прошла валидацию"
                            ),
                        )
                    )
                    continue

                code_key = (preview_row.nomenclature_code or "").strip().upper()
                nomenclature_row = nomenclature_by_code.get(code_key)
                if (
                    nomenclature_row is None
                    or preview_row.as_of_date is None
                    or preview_row.available_qty is None
                ):
                    skipped_count += 1
                    error_count += 1
                    commit_rows.append(
                        InventoryBalanceImportCommitRow(
                            row_no=preview_row.row_no,
                            as_of_date=preview_row.as_of_date,
                            nomenclature_code=preview_row.nomenclature_code,
                            status="error",
                            message="Строка не прошла повторную проверку",
                        )
                    )
                    continue

                inserted = upsert_inventory_balance_row(
                    cursor,
                    as_of_date=preview_row.as_of_date,
                    nomenclature_id=nomenclature_row["nomenclature_id"],
                    available_qty=preview_row.available_qty,
                )
                if inserted:
                    created_count += 1
                    commit_rows.append(
                        InventoryBalanceImportCommitRow(
                            row_no=preview_row.row_no,
                            as_of_date=preview_row.as_of_date,
                            nomenclature_code=preview_row.nomenclature_code,
                            status="created",
                            message="Создано",
                        )
                    )
                else:
                    updated_count += 1
                    commit_rows.append(
                        InventoryBalanceImportCommitRow(
                            row_no=preview_row.row_no,
                            as_of_date=preview_row.as_of_date,
                            nomenclature_code=preview_row.nomenclature_code,
                            status="updated",
                            message="Обновлено",
                        )
                    )

        connection.commit()
        return InventoryBalanceImportCommitResponse(
            import_mode=IMPORT_MODE_UPSERT,
            total_rows=preview_response.total_rows,
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=skipped_count,
            error_count=error_count,
            rows=commit_rows,
        )
    except psycopg2.Error as exc:
        if connection is not None:
            connection.rollback()

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Не удалось выполнить импорт остатков.",
        ) from exc
    finally:
        if connection is not None:
            connection.close()
